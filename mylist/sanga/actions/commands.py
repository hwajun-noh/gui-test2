# commands.py - sanga 액션 명령 모듈
import os
import sys
import glob
import time
import json
from datetime import datetime
import logging
import traceback
import requests

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt, QModelIndex, QUrl, pyqtSlot, QObject, pyqtSignal
from PyQt5.QtGui import QStandardItem, QBrush, QPixmap, QIcon
from PyQt5.QtWidgets import QMessageBox, QInputDialog, QFileDialog, QApplication

# 외부 엑셀 모듈
import openpyxl
from openpyxl import Workbook

# 상수 및 UI 다이얼로그
from mylist.constants import PENDING_COLOR, RE_AD_BG_COLOR, NEW_AD_BG_COLOR
from dialogs import NaverShopSearchDialog

# 로거 인스턴스
logger = logging.getLogger(__name__)

class SangaCommands(QObject):
    """상가 데이터 명령 처리 클래스"""
    
    # 명령 실행 결과 시그널
    commandExecuted = pyqtSignal(str, bool)  # 명령, 성공여부
    
    def __init__(self, model=None, pending_manager=None, parent=None):
        """
        초기화
        
        Args:
            model: SangaModel 인스턴스
            pending_manager: 변경 사항 관리자
            parent: 부모 객체
        """
        super().__init__(parent)
        self.logger = logging.getLogger(__name__)
        self.model = model
        self.pending_manager = pending_manager
        
    def add_row(self, initial_data=None):
        """
        새 행 추가
        
        Args:
            initial_data (dict, optional): 초기 데이터
            
        Returns:
            int: 추가된 행의 인덱스
        """
        if self.model is None:
            self.logger.error("모델이 설정되지 않았습니다.")
            self.commandExecuted.emit("add_row", False)
            return -1
            
        try:
            # 임시 ID 생성
            temp_id = self._generate_temp_id() if self.pending_manager else -1
            
            # 초기 데이터가 있으면 사용, 없으면 기본값 설정
            if initial_data is None:
                initial_data = {
                    "address": "",
                    "manager": "",
                    "property_no": "",
                    "room_no": "",
                    "floor": "",
                    "move_in_date": "",
                    "monthly": "",
                    "deposit": "",
                    "maintenance_fee": "",
                    "key_money": "",
                    "rooms_bath": "",
                    "area": "",
                    "current_business": "",
                    "business_restriction": "",
                    "parking": "",
                    "building_type": "",
                    "private_entrance": "",
                    "phone": "",
                    "memo": "",
                    "approval_date": "",
                    "re_ad_yn": "N"
                }
                
            # temp_id 추가
            if temp_id:
                initial_data["id"] = temp_id
                
            # 행 추가
            row_idx = self.model.add_row(initial_data)
            
            # 변경 관리자에 추가
            if self.pending_manager:
                self.pending_manager.add_pending_shop_addition(temp_id, row_idx)
                
            self.commandExecuted.emit("add_row", True)
            return row_idx
            
        except Exception as e:
            self.logger.error(f"행 추가 중 오류 발생: {e}", exc_info=True)
            self.commandExecuted.emit("add_row", False)
            return -1
    
    def delete_rows(self, row_indices):
        """
        행 삭제
        
        Args:
            row_indices (list): 삭제할 행 인덱스 목록
            
        Returns:
            int: 삭제된 행 수
        """
        if self.model is None:
            self.logger.error("모델이 설정되지 않았습니다.")
            self.commandExecuted.emit("delete_rows", False)
            return 0
            
        try:
            # 모델의 delete_rows 메서드 호출
            self.model.delete_rows(row_indices)
            deleted_count = len(row_indices)
            
            self.commandExecuted.emit("delete_rows", deleted_count > 0)
            return deleted_count
            
        except Exception as e:
            self.logger.error(f"행 삭제 중 오류 발생: {e}", exc_info=True)
            self.commandExecuted.emit("delete_rows", False)
            return 0
    
    def update_model_row(self, model, row_idx, headers, db_row_data):
        """
        모델의 한 행을 업데이트하는 메소드
        
        Args:
            model: 업데이트할 QStandardItemModel
            row_idx: 업데이트할 행 인덱스
            headers: 컬럼 헤더 목록
            db_row_data: 데이터베이스 행 데이터
            
        Returns:
            bool: 성공 여부
        """
        try:
            # 실행 시작 로깅
            self.logger.debug(f"update_model_row 시작: 행={row_idx}, 데이터 키={list(db_row_data.keys() if db_row_data else [])}")
            
            # 기본 확인
            if not model or row_idx < 0 or row_idx >= model.rowCount():
                self.logger.error(f"유효하지 않은 모델 또는 행 인덱스: {row_idx}")
                return False
            
            # DB 필드명과 UI 필드명 매핑 (model.column_map으로부터 가져올 수 있음)
            column_map = {}
            if hasattr(self.model, 'column_map'):
                column_map = self.model.column_map
            
            for col_idx, header_name in enumerate(headers):
                # DB 필드명 조회
                db_key = column_map.get(header_name)
                # DB 값 조회
                raw_value = db_row_data.get(db_key) if db_key else None
                
                # 헤더별 특수 처리
                if header_name == "주소":
                    cell_val = f"{db_row_data.get('dong', '')} {db_row_data.get('jibun', '')}".strip()
                elif header_name == "층":
                    cell_val = f"{db_row_data.get('curr_floor', 0)}/{db_row_data.get('total_floor', 0)}"
                elif header_name == "보증금/월세":
                    cell_val = f"{db_row_data.get('deposit', 0)}/{db_row_data.get('monthly', 0)}"
                elif header_name == "보증금(만원)":
                    cell_val = str(db_row_data.get('deposit', ''))
                elif header_name == "월세(만원)":
                    cell_val = str(db_row_data.get('monthly', ''))
                elif header_name == "방/화장실":
                    r = db_row_data.get("rooms", "")
                    b = db_row_data.get("baths", "")
                    cell_val = f"방{r}/{b}" if r or b else ""
                elif header_name == "평수(공급/전용)":
                    cell_val = str(db_row_data.get('area', ''))
                elif header_name == "매물번호":
                    cell_val = f"{db_row_data.get('naver_property_no', '')}/{db_row_data.get('serve_property_no', '')}"
                elif header_name == "재광고":
                    cell_val = "재광고" if db_row_data.get("re_ad_yn", "N") == "Y" else "새광고"
                else:
                    cell_val = str(raw_value) if raw_value is not None else ""
                
                # 아이템 생성
                item = QStandardItem(str(cell_val))
                
                # 주소 열 추가 처리
                if header_name == "주소":
                    # 사진 경로 처리
                    folder_path = db_row_data.get("photo_path", "") or ""
                    rep_img_path = ""
                    
                    if folder_path and os.path.isdir(folder_path):
                        try:
                            files = [f for f in os.listdir(folder_path) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.gif'))]
                            if files: 
                                rep_img_path = os.path.join(folder_path, sorted(files)[0])
                        except OSError as e: 
                            self.logger.warning(f"Cannot access folder path '{folder_path}': {e}")
                    
                    # 대표 이미지가 있으면 아이콘 설정
                    if rep_img_path and os.path.isfile(rep_img_path):
                        try:
                            pixmap = QPixmap(rep_img_path).scaled(24, 24, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                            if not pixmap.isNull():
                                item.setIcon(QIcon(pixmap))
                                # 백슬래시 문제 해결: f-string 대신 일반 문자열 연결 사용
                                file_url = "file:///" + rep_img_path.replace('\\', '/')
                                item.setToolTip(f'<img src="{file_url}" width="200">')
                            else: 
                                item.setToolTip(cell_val)
                        except Exception as img_err:
                            self.logger.warning(f"Error creating shop icon/tooltip for {rep_img_path}: {img_err}")
                            item.setToolTip(cell_val)
                    else: 
                        item.setToolTip(cell_val)
                    
                    # 메타데이터 설정
                    item.setData(folder_path, Qt.UserRole + 10)  # 사진 폴더 경로
                    item.setData(rep_img_path, Qt.UserRole + 11)  # 대표 이미지 경로
                    item.setData(db_row_data.get("status_cd", ""), Qt.UserRole + 1)  # 상태 코드
                    item.setData(db_row_data.get("id"), Qt.UserRole + 3)  # ID
                
                # 아이템 설정
                model.setItem(row_idx, col_idx, item)
                
            # 재광고 여부에 따른 배경색 설정
            re_ad_yn = db_row_data.get("re_ad_yn", "N") == "Y"
            row_bg = RE_AD_BG_COLOR if re_ad_yn else NEW_AD_BG_COLOR
            
            for c in range(model.columnCount()):
                cell = model.item(row_idx, c)
                if cell:
                    cell.setBackground(row_bg)
            
            self.logger.debug(f"update_model_row 완료: 행={row_idx}")
            self.commandExecuted.emit("update_model_row", True)
            return True
            
        except Exception as e:
            self.logger.error(f"행 업데이트 중 오류 발생: {e}", exc_info=True)
            self.commandExecuted.emit("update_model_row", False)
            return False
    
    def update_cell(self, row_idx, col_idx, value):
        """
        셀 값 업데이트
        
        Args:
            row_idx (int): 행 인덱스
            col_idx (int): 열 인덱스
            value (str): 새 값
            
        Returns:
            bool: 성공 여부
        """
        if self.model is None:
            self.logger.error("모델이 설정되지 않았습니다.")
            self.commandExecuted.emit("update_cell", False)
            return False
            
        try:
            # 현재 셀 정보
            model = self.model.get_model()
            model_item = model.item(row_idx, col_idx)
            if not model_item:
                self.logger.error(f"셀({row_idx}, {col_idx})을 찾을 수 없습니다.")
                return False
                
            old_value = model_item.text()
            if old_value == value:
                self.logger.debug(f"셀({row_idx}, {col_idx}) 값이 변경되지 않았습니다: {value}")
                return True
                
            # 값 업데이트
            model_item.setText(value)
            
            # 변경사항 등록
            if self.pending_manager:
                row_data = self.model.get_row_data(row_idx)
                record_id = row_data.get("id")
                
                if record_id > 0:  # 실제 DB 행
                    # 변경된 필드 결정
                    headers = [model.horizontalHeaderItem(c).text() for c in range(model.columnCount())]
                    header = headers[col_idx]
                    db_field = self.model.column_map.get(header)
                    
                    if db_field:
                        update_payload = {
                            "id": record_id,
                            db_field: value
                        }
                        self.pending_manager.add_pending_shop_update(update_payload)
                        
                        # 배경색 변경
                        model_item.setBackground(PENDING_COLOR)
                        
            self.commandExecuted.emit("update_cell", True)
            return True
            
        except Exception as e:
            self.logger.error(f"셀 업데이트 중 오류 발생: {e}", exc_info=True)
            self.commandExecuted.emit("update_cell", False)
            return False
    
    def change_manager(self, row_indices, new_manager):
        """
        담당자 변경
        
        Args:
            row_indices (list): 변경할 행 인덱스 목록
            new_manager (str): 새 담당자
            
        Returns:
            int: 변경된 행 수
        """
        if self.model is None:
            self.logger.error("모델이 설정되지 않았습니다.")
            self.commandExecuted.emit("change_manager", False)
            return 0
            
        try:
            # 담당자 열 인덱스 찾기
            model = self.model.get_model()
            headers = [model.horizontalHeaderItem(c).text() for c in range(model.columnCount())]
            manager_col_idx = -1
            for idx, header in enumerate(headers):
                if header == "담당자":
                    manager_col_idx = idx
                    break
                    
            if manager_col_idx == -1:
                self.logger.error("담당자 열을 찾을 수 없습니다.")
                return 0
                
            # 담당자 변경
            updated_count = 0
            for row_idx in row_indices:
                # 셀 업데이트
                if self.update_cell(row_idx, manager_col_idx, new_manager):
                    updated_count += 1
                    
            self.commandExecuted.emit("change_manager", updated_count > 0)
            return updated_count
            
        except Exception as e:
            self.logger.error(f"담당자 변경 중 오류 발생: {e}", exc_info=True)
            self.commandExecuted.emit("change_manager", False)
            return 0
    
    def change_re_ad(self, row_indices, is_re_ad):
        """
        재광고 여부 변경
        
        Args:
            row_indices (list): 변경할 행 인덱스 목록
            is_re_ad (bool): 재광고 여부
            
        Returns:
            int: 변경된 행 수
        """
        if self.model is None:
            self.logger.error("모델이 설정되지 않았습니다.")
            self.commandExecuted.emit("change_re_ad", False)
            return 0
            
        try:
            # 재광고 열 인덱스 찾기
            model = self.model.get_model()
            headers = [model.horizontalHeaderItem(c).text() for c in range(model.columnCount())]
            re_ad_col_idx = -1
            for idx, header in enumerate(headers):
                if header == "재광고":
                    re_ad_col_idx = idx
                    break
                    
            if re_ad_col_idx == -1:
                self.logger.error("재광고 열을 찾을 수 없습니다.")
                return 0
                
            # 재광고 여부 변경
            updated_count = 0
            re_ad_text = "재광고" if is_re_ad else "새광고"
            
            for row_idx in row_indices:
                # 셀 업데이트
                if self.update_cell(row_idx, re_ad_col_idx, re_ad_text):
                    updated_count += 1
                    
            self.commandExecuted.emit("change_re_ad", updated_count > 0)
            return updated_count
            
        except Exception as e:
            self.logger.error(f"재광고 여부 변경 중 오류 발생: {e}", exc_info=True)
            self.commandExecuted.emit("change_re_ad", False)
            return 0
                
    def _generate_temp_id(self):
        """임시 ID 생성 (pending_manager 사용)"""
        if self.pending_manager and hasattr(self.pending_manager, "generate_temp_id"):
            return self.pending_manager.generate_temp_id()
        # 백업 메서드: 현재 시간 기반 음수 ID 생성
        import random
        return -1000 - random.randint(0, 9000)

def on_save_mylist_shop_changes(logic_instance):
    """
    상가 변경사항 저장 요청을 save_handler로 전달합니다.
    """
    if logic_instance.container and logic_instance.container.save_handler:
        logger.info("Manual save requested for shop tab. Forwarding to save_handler.")
        logic_instance.container.save_handler.save_pending_shop_changes()
    else:
         logger.error("Cannot save shop changes: container or save_handler is missing.")
         QMessageBox.warning(logic_instance.parent_app, "오류", "저장 핸들러를 찾을 수 없어 저장할 수 없습니다.")

def on_naver_search_clicked(logic_instance):
    """
    네이버 매물 검색 다이얼로그를 엽니다.
    """
    logger.info("MyListSanga: 'Naver Property Search' button clicked.")

    # 다이얼로그 재사용 또는 새로 생성
    if hasattr(logic_instance, 'naver_search_dialog') and logic_instance.naver_search_dialog and logic_instance.naver_search_dialog.isVisible():
        logger.info("Existing Naver Search Dialog found for MyListSanga. Activating.")
        logic_instance.naver_search_dialog.activateWindow()
        logic_instance.naver_search_dialog.raise_()
        return

    # 다이얼로그 인스턴스를 logic_instance 속성으로 저장
    logic_instance.naver_search_dialog = NaverShopSearchDialog(
        parent_app=logic_instance.parent_app,
        mylist_tab=logic_instance.container 
    )

    # 시그널 연결
    if hasattr(logic_instance.parent_app, 'all_tab') and hasattr(logic_instance.parent_app.all_tab, 'search_by_address'):
        # 메인 앱의 중앙 핸들러에 연결
        if hasattr(logic_instance.parent_app, 'handle_address_selection_from_dialog'):
            logic_instance.naver_search_dialog.addressClicked.connect(logic_instance.parent_app.handle_address_selection_from_dialog)
            logger.info("Connected NaverShopSearchDialog.addressClicked to main_app.handle_address_selection_from_dialog")
        else:
            logger.warning("Could not connect addressClicked signal: main_app.handle_address_selection_from_dialog slot not found.")
    else:
        logger.warning("Could not connect addressClicked signal: all_tab or search_by_address slot not found.")

    logic_instance.naver_search_dialog.show()
    logger.info("Naver Search Dialog shown (modeless).")

def add_shop_row_with_data(logic_instance, row_dict_from_naver):
    """
    네이버 검색에서 가져온 데이터로 새 상가 행을 추가합니다.
    """
    if not row_dict_from_naver:
        return
    
    # 컨테이너 메서드를 호출하여 행 추가, 네이버 형식 파싱 지정
    if logic_instance.container:
        logic_instance.container.add_new_shop_row(initial_data=row_dict_from_naver, parse_naver_format=True)

def filter_table_by_address(logic_instance, address_str: str):
    """
    주어진 주소 문자열로 상가 테이블을 필터링합니다.
    """
    if not address_str:
        return  # 필터링할 내용 없음
    
    address_str = address_str.strip().lower()  # 대소문자 구분 없음
    
    model = logic_instance.mylist_shop_model
    view = logic_instance.mylist_shop_view
    
    if not model or not view:
        return
    
    # 주소 열 인덱스 확인
    headers = [model.horizontalHeaderItem(c).text() for c in range(model.columnCount())]
    try:
        addr_col_idx = headers.index("주소")
    except ValueError:
        logger.warning("Cannot filter by address: '주소' column not found")
        return
    
    # 간단한 필터링: 셀 텍스트에 address_str이 포함되어 있는지 확인
    # 더 정교한 매칭으로 개선 가능
    for row in range(model.rowCount()):
        item = model.item(row, addr_col_idx)
        if not item:
            continue
        
        cell_text = item.text().lower()
        row_visible = address_str in cell_text
        
        # 필터 일치 여부에 따라 행 숨기기/표시
        view.setRowHidden(row, not row_visible)
    
    matching_count = sum(1 for row in range(model.rowCount()) if not view.isRowHidden(row))
    logger.info(f"Address filter applied: '{address_str}' - {matching_count} rows visible")

def copy_mylist_shop_row(logic_instance, source_row_idx):
    """
    행을 복사하여 새 임시 행으로 추가합니다.
    """
    model = logic_instance.mylist_shop_model
    if not model or source_row_idx < 0 or source_row_idx >= model.rowCount():
        return

    col_count = model.columnCount()
    copied_values = [model.item(source_row_idx, c).text() if model.item(source_row_idx, c) else "" for c in range(col_count)]

    # 컨테이너 메서드를 사용하여 행 추가 및 pending 상태 & 임시 ID 처리
    if logic_instance.container:
        logic_instance.container.add_new_shop_row(initial_data=copied_values)

def export_selected_shop_to_excel(logic_instance):
    """
    선택한 행을 사전 정의된 매핑에 따라 Excel 파일로 내보냅니다.
    """
    view = logic_instance.mylist_shop_view
    model = logic_instance.mylist_shop_model

    if not view or not model:
        logger.warning("View 또는 Model을 찾을 수 없습니다.")
        return

    # 1) selection: 단일 셀만 클릭해도 해당 row 포함
    selection_model = view.selectionModel()
    if not selection_model:
        logger.warning("Selection Model을 찾을 수 없습니다.")
        return

    selected_indexes = selection_model.selectedIndexes()
    if not selected_indexes:
        logger.warning("선택된 셀이 없습니다.")
        return

    # 선택된 셀들의 row 인덱스 -> 중복 제거 -> 오름차순 정렬
    row_set = set(idx.row() for idx in selected_indexes)
    row_indexes = sorted(row_set)
    if not row_indexes:
        logger.warning("선택된 행이 없습니다.")
        return

    # (A) 엑셀 열 매핑: (엑셀열 인덱스, "엑셀헤더명", mylist col 인덱스)
    excel_map = [
        (0, "주소", 0),
        (1, "호", 1),
        (2, "권리금", 5),
        (3, "현업종", 6),
        (4, "보증금/월세", 3),
        (5, "관리비(만원)", 4),
        (6, "평수(㎡)", 7),
        (7, "해당층/총층", 2),
        (8, "담당자", 10),
        (9, "연락처", 8),
        (10, "방/화장실", 15),
        (11, "주차대수", 12),
        (12, "매물번호", 9),
        (13, "용도", 13),
        (14, "사용승인일", 14), # 하이픈 제거 필요
        (15, "소유자명", 18),
        (16, "관계", 19),
        (17, "사진경로", 17),
        (18, "광고종료일", 16),
        (19, "메모", 11),
    ]

    # '재광고' 상태를 가져올 모델 컬럼 인덱스 확인
    all_headers = [model.horizontalHeaderItem(c).text() for c in range(model.columnCount())]
    try:
        re_ad_col_index = all_headers.index("재광고")
    except ValueError:
        logger.warning("'재광고' 컬럼을 찾을 수 없어 담당자명 처리가 제한될 수 있습니다.")
        re_ad_col_index = -1

    # (B) Workbook 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "마이리스트_상가"

    # (C) 엑셀 첫 행(1행) => 헤더 작성 (excel_map 기준)
    for excel_col_idx, header_text, _ in excel_map:
        ws.cell(row=1, column=excel_col_idx + 1).value = header_text

    # (D) 엑셀 row=2부터 실제 mylist_shop 데이터 (excel_map 기준)
    start_excel_row = 2
    for i, row_idx in enumerate(row_indexes):
        excel_row = start_excel_row + i

        # 재광고 상태 값 가져오기 (담당자 처리용)
        re_ad_val = ""
        if re_ad_col_index != -1:
            re_ad_item = model.item(row_idx, re_ad_col_index)
            re_ad_val = re_ad_item.text().strip() if re_ad_item else ""

        # excel_map 순서대로 값 가져와서 변환 후 대입
        for excel_col_idx, _, mylist_col_idx in excel_map:
            item_ = model.item(row_idx, mylist_col_idx)
            text_val = item_.text() if item_ else ""

            # 데이터 변환 로직 적용
            # 권리금(col=5): "무권리" -> ""
            if mylist_col_idx == 5 and text_val.strip() == "무권리":
                text_val = ""
            # 현업종(col=6): "공실" -> ""
            elif mylist_col_idx == 6 and text_val.strip() == "공실":
                text_val = ""
            # 사용승인일(col=14): 하이픈 제거
            elif mylist_col_idx == 14:
                text_val = text_val.replace("-", "")
            # 담당자(col=10): 재광고 상태에 따라 형식 변경
            elif mylist_col_idx == 10:
                manager_name = text_val.strip()
                if manager_name:
                    if re_ad_val == "재광고":
                        text_val = f"{manager_name}(재광고)"
                    else:
                        text_val = manager_name
                else:
                    text_val = ""

            ws.cell(row=excel_row, column=excel_col_idx + 1).value = text_val

    # (E) 파일명 자동 생성 및 저장
    try:
        current_manager_name = getattr(logic_instance, 'current_manager', 'UnknownManager')
        now_str = datetime.now().strftime("%Y_%m_%d_%H%M")
        manager_str = current_manager_name.replace(" ", "_")
        filename = f"{now_str}_{manager_str}.xlsx"

        wb.save(filename)
        logger.info(f"엑셀 다운로드 완료 ({len(row_indexes)} 행) => {filename}")

    except AttributeError:
        logger.error("현재 매니저 이름 ('current_manager')을 가져올 수 없어 파일명 생성이 제한됩니다.")
        # 파일 저장 시도 (기본 이름 사용)
        try:
            fallback_filename = f"마이리스트_상가_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            wb.save(fallback_filename)
            logger.info(f"엑셀 다운로드 완료 ({len(row_indexes)} 행) => {fallback_filename}")
        except Exception as e_save:
            logger.error(f"엑셀 파일 저장 중 오류 발생: {e_save}")
    except Exception as e:
        logger.error(f"엑셀 파일 저장 중 오류 발생: {e}")

def on_open_sanga_tk_for_mylist_shop(logic_instance):
    """
    네이버 부동산 검수용 TKinter 창을 엽니다.
    """
    try:
        # 현재 선택된 행 확인
        view = logic_instance.mylist_shop_view
        if not view or view.currentIndex().row() < 0:
            QMessageBox.information(
                logic_instance.parent_app,
                "선택 없음",
                "검수할 행을 선택하세요."
            )
            return
        
        # 컨테이너의 네이버부동산 검수 실행 함수 호출
        if logic_instance.container and hasattr(logic_instance.container, 'launch_naver_check_for_mylist'):
            logic_instance.container.launch_naver_check_for_mylist()
        else:
            # 컨테이너가 없거나 메서드가 없는 경우
            logger.error("MyListContainer 또는 launch_naver_check_for_mylist 메서드를 찾을 수 없습니다.")
            QMessageBox.critical(
                logic_instance.parent_app,
                "실행 오류",
                "네이버부동산 검수 기능을 실행할 수 없습니다. MyListContainer 구성을 확인하세요."
            )
            
    except Exception as e:
        logger.error(f"네이버부동산 검수 실행 중 오류: {e}")
        traceback.print_exc()
        
        QMessageBox.critical(
            logic_instance.parent_app,
            "검수 창 실행 오류",
            f"네이버부동산 검수 창을 실행하는 중 오류가 발생했습니다:\n{e}"
        )

def highlight_mylist_shop_row_by_id(logic_instance, pk_id):
    """
    ID로 행을 찾아 강조 표시합니다.
    """
    row_index = logic_instance.find_mylist_shop_row_by_id(pk_id)
    view = logic_instance.mylist_shop_view
    
    if row_index is not None and view:
        view.selectRow(row_index)
        view.scrollTo(view.model().index(row_index, 0))
        logger.info(f"Highlighted row {row_index} with ID {pk_id}")
    else:
        logger.warning(f"Could not find row with ID {pk_id} to highlight")

def _handle_save_complete(logic_instance, result):
    """
    (메인 스레드 슬롯) 저장 작업 완료를 처리합니다.
    """
    logger.info(f"_handle_save_complete: Received save result (status={result.get('status') if result else 'None'}).")

    if not result:
        logger.error("_handle_save_complete: Received empty result.")
        QMessageBox.critical(logic_instance.parent_app, "저장 실패", "서버로부터 응답이 없습니다.")
        logic_instance.container.update_autosave_status("저장 실패 (응답 없음)")
        return

    if result.get("status") == "ok":
        logger.info("_handle_save_complete: Save successful. Processing updates.")
        added_results = result.get("added_ids", {})  # {temp_id: new_db_id}
        updated_count = result.get("updated_count", 0)
        deleted_count = result.get("deleted_count", 0)

        # 추가된 항목 처리: 임시 ID를 실제 DB ID로 업데이트
        model = logic_instance.mylist_shop_model
        if model:
            for temp_id_str, new_id in added_results.items():
                try:
                    temp_id = int(temp_id_str)
                    logger.debug(f"Updating temp_id {temp_id} to new_id {new_id}")
                    logic_instance._update_mylist_shop_row_id(temp_id, new_id)
                    # 새로 추가된 행의 배경색도 초기화
                    row_idx = logic_instance.find_mylist_shop_row_by_id(new_id)
                    if row_idx is not None:
                        for col in range(model.columnCount()):
                            item = model.item(row_idx, col)
                            if item:
                                item.setBackground(QBrush(Qt.NoBrush))  # 배경색 초기화
                except ValueError:
                    logger.error(f"_handle_save_complete: Invalid temp_id format '{temp_id_str}' received.")
                except Exception as e_update:
                    logger.error(f"_handle_save_complete: Error updating row ID for temp_id {temp_id_str}: {e_update}", exc_info=True)

            # 저장 성공한 업데이트 항목 배경색 복원
            # 저장 *전*의 pending 상태를 가져와야 함
            pending_updates_before_save = logic_instance.container.pending_manager.get_pending_shop_changes().get("updated", [])
            for update_item in pending_updates_before_save:
                record_id = update_item.get("id")
                if record_id is None: continue  # ID 없으면 스킵

                row_idx = logic_instance.find_mylist_shop_row_by_id(record_id)
                if row_idx is not None:
                    # update_item 딕셔너리에 포함된 모든 필드에 대해 배경색 복원 시도
                    for db_field_name in update_item.keys():
                        if db_field_name == "id": continue  # ID 필드는 스킵

                        display_col_name = None
                        # DB 필드명 -> 표시 열 이름 변환 (주의: 복합 필드는 이 로직으로 처리 안됨)
                        for disp_name, db_name in logic_instance.parent_app.COLUMN_MAP_MYLIST_SHOP_DISPLAY_TO_DB.items():
                            if db_name == db_field_name:
                                display_col_name = disp_name
                                break
                        
                        if display_col_name:
                            try:
                                headers = [model.horizontalHeaderItem(c).text() for c in range(model.columnCount())]
                                col_idx = headers.index(display_col_name)
                                item = model.item(row_idx, col_idx)
                                if item and item.background() != QBrush(Qt.NoBrush):  # 이미 초기화된 경우는 제외
                                    item.setBackground(QBrush(Qt.NoBrush))  # 배경색 초기화
                                    logger.debug(f"Reset background for updated item: ID={record_id}, Row={row_idx}, Col={col_idx} ({display_col_name})")
                            except ValueError:
                                logger.warning(f"Could not find column index for display name '{display_col_name}' (DB field: {db_field_name}) while resetting background.")
                            except Exception as e_reset:
                                logger.error(f"Error resetting background for item (ID:{record_id}, Row:{row_idx}, Col:{display_col_name}): {e_reset}")

        # 성공적으로 저장된 항목에 대해 pending 상태 클리어
        logic_instance.container.pending_manager.clear_shop_pending_state()

        status_msg = f"저장 완료 (추가: {len(added_results)}, 수정: {updated_count}, 삭제: {deleted_count})"
        QMessageBox.information(logic_instance.parent_app, "저장 완료", status_msg)
        logic_instance.container.update_autosave_status(status_msg)

    else:
        error_msg = result.get("message", "알 수 없는 서버 오류")
        logger.error(f"_handle_save_complete: Save failed. Server message: {error_msg}")
        QMessageBox.critical(logic_instance.parent_app, "저장 실패", f"서버 오류:\n{error_msg}")
        logic_instance.container.update_autosave_status(f"저장 실패 ({error_msg[:30]}...)")

def run_naver_inspect_app(AppClass, params):
    """내부 함수: NaverSangaInspectApp 실행 및 예외 처리"""
    try:
        # SangaCheckAppMylist.run()은 data_list를 반환
        result_data = AppClass(**params).run()
        logger.info(f"검수 작업 완료: {len(result_data)}개 항목 처리됨")
        return result_data
    except Exception as run_e:
        logger.error(f"검수 창 실행 오류: {run_e}")
        traceback.print_exc()
        return None 