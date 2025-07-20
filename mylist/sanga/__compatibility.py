"""
마이리스트 상가 호환성 레이어

기존 MyListSangaLogic 클래스와 새 모듈식 구조 사이의 호환성 레이어 제공
"""
import logging
from PyQt5.QtCore import Qt, QTimer, QObject
from PyQt5.QtWidgets import QWidget, QVBoxLayout

# 기존 코드 임포트
from mylist_sanga_logic import MyListSangaLogic

# 새 모듈 임포트 (직접 모듈에서 임포트)
from mylist.sanga.data.models import SangaModel
from mylist.sanga.data.loaders import SangaDataLoader
from mylist.sanga.events.view_events import SangaViewEvents
from mylist.sanga.events.item_events import SangaItemEvents
from mylist.sanga.events.context_menu_events import SangaContextMenuEvents
from mylist.sanga.ui.components import init_sanga_ui
from mylist.sanga.actions.commands import SangaCommands

class SangaBridge(QObject):
    """
    기존 MyListSangaLogic과 새 모듈식 구조 사이의 브릿지
    
    이 클래스는 이행 기간 동안 사용되며, 모듈화가 완료되면 제거될 수 있습니다.
    """
    
    def __init__(self, parent_app=None, container=None):
        """
        초기화
        
        Args:
            parent_app: 부모 애플리케이션
            container: 컨테이너 객체
        """
        super().__init__(parent=container)  # QObject 초기화, container를 부모로 설정
        
        self.logger = logging.getLogger(__name__)
        self.parent_app = parent_app
        self.container = container
        
        # 서버 정보
        self.server_host = getattr(container, 'server_host', 'localhost')
        self.server_port = getattr(container, 'server_port', 8000)
        
        # 기존 로직 클래스 (레거시)
        self.legacy_logic = MyListSangaLogic(parent_app, container)
        
        # 새 모듈 컴포넌트 (모듈식)
        self.model = SangaModel()
        self.loader = SangaDataLoader(logic_instance=self)
        self.commands = SangaCommands(self.model, getattr(container, 'pending_manager', None))
        
        # 이벤트 핸들러
        self.view_events = SangaViewEvents(self)
        self.item_events = SangaItemEvents(self)
        self.context_menu_events = SangaContextMenuEvents(self)
        
        # 내부 상태
        self.using_legacy = True  # 레거시 모드 사용 (호환성 문제로 되돌림)
        
        # 호환성을 위한 속성 추가
        self.mylist_shop_model = self.legacy_logic.mylist_shop_model if self.using_legacy else self.model.model
        self.mylist_shop_view = None  # UI 초기화 후 설정됨
        
        # row_manager에서 사용하는 속성 추가
        self.manager_summary_label = None  # UI 초기화 시 설정될 예정
        
        # 로깅 설정
        self.logger.info("SangaBridge 인스턴스 생성됨")
        
    # 프로퍼티 메서드 추가 - 호출 시점에 항상 올바른 값 반환
    @property
    def mylist_shop_model(self):
        """모델 참조를 동적으로 가져오는 프로퍼티"""
        if self.using_legacy:
            if hasattr(self.legacy_logic, 'mylist_shop_model'):
                return self.legacy_logic.mylist_shop_model
        else:
            return self.model.model
        return None
        
    @mylist_shop_model.setter
    def mylist_shop_model(self, value):
        """모델 참조를 설정하는 setter (호환성용)"""
        if not self.using_legacy:
            self.model.model = value
        # 레거시 모드에서는 무시 (레거시 로직이 직접 관리)
            
    @property
    def manager_summary_label(self):
        """요약 레이블 참조를 동적으로 가져오는 프로퍼티"""
        if self.using_legacy:
            if hasattr(self.legacy_logic, 'manager_summary_label'):
                return self.legacy_logic.manager_summary_label
        else:
            # 모듈식 모드에서의 레이블 가져오기
            if hasattr(self, '_manager_summary_label'):
                return self._manager_summary_label
        return None
        
    @manager_summary_label.setter
    def manager_summary_label(self, value):
        """요약 레이블 참조를 설정하는 setter"""
        if not self.using_legacy:
            self._manager_summary_label = value
        # 레거시 모드에서는 무시 (레거시 로직이 직접 관리)
        
    def init_ui(self):
        """
        UI 초기화 (기존 MyListSangaLogic.init_ui와 호환)
        
        Returns:
            QWidget: 상가 탭 위젯
        """
        if self.using_legacy:
            widget = self.legacy_logic.init_ui()
            # 이제 직접 속성 할당이 필요 없음 (프로퍼티가 자동으로 처리)
            # 뷰 참조는 여전히 필요 (프로퍼티 없음)
            if hasattr(self.legacy_logic, 'mylist_shop_view'):
                self._mylist_shop_view = self.legacy_logic.mylist_shop_view
            self.logger.info(f"Legacy 모드: UI 초기화 완료")
            return widget
        else:
            # 새 UI 생성 및 설정
            widget = init_sanga_ui(self)
            
            # 호환성 속성 업데이트 (뷰만 필요)
            self._mylist_shop_view = getattr(self, 'mylist_shop_view', None)
            
            # 뷰 이벤트 설정
            if hasattr(self, '_mylist_shop_view') and self._mylist_shop_view:
                self.view_events.setup_view_signals(self._mylist_shop_view)
                self.context_menu_events.setup_context_menu(self._mylist_shop_view)
            
            return widget

    @property
    def mylist_shop_view(self):
        """뷰 참조를 가져오는 프로퍼티"""
        if self.using_legacy:
            if hasattr(self.legacy_logic, 'mylist_shop_view'):
                return self.legacy_logic.mylist_shop_view
        if hasattr(self, '_mylist_shop_view'):
            return self._mylist_shop_view
        return None
    
    @mylist_shop_view.setter
    def mylist_shop_view(self, value):
        """뷰 참조를 설정하는 setter"""
        self._mylist_shop_view = value
            
    def load_data(self, manager=None, filters=None):
        """
        데이터 로드 (기존 MyListSangaLogic.load_data와 호환)
        
        Args:
            manager: 담당자 필터 (없으면 전체)
            filters: 추가 필터 조건
        """
        if self.using_legacy:
            # 기존 로직은 인자가 1개만 있기 때문에 직접 호출
            return self.legacy_logic.load_data()
        else:
            # 현재 관리자와 역할 정보 가져오기
            current_manager = getattr(self.container, 'current_manager', 'admin')
            role = getattr(self.container, 'current_role', 'admin')
            
            # 로더를 통해 데이터 로드
            self.loader.load_data(manager or current_manager, filters)
    
    def _update_model_row(self, model, row_idx, headers, db_data):
        """
        모델 행 업데이트 (직접 호출용)
        
        Args:
            model: 업데이트할 모델
            row_idx: 행 인덱스
            headers: 헤더 목록
            db_data: DB 데이터
        """
        self.logger.info(f"SangaBridge._update_model_row 직접 호출됨: row_idx={row_idx}")
        if self.using_legacy:
            if hasattr(self.legacy_logic, '_update_model_row'):
                self.logger.info("레거시 로직의 _update_model_row 메서드 호출")
                return self.legacy_logic._update_model_row(model, row_idx, headers, db_data)
            else:
                self.logger.error("레거시 로직에 _update_model_row 메서드가 없음")
                return False
        else:
            # 모듈식 구현에서는 commands를 통해 처리
            try:
                self.logger.info("모듈식 로직의 update_model_row 메서드 호출")
                self.commands.update_model_row(model, row_idx, headers, db_data)
                return True
            except Exception as e:
                self.logger.error(f"모듈식 구현의 update_model_row 호출 중 오류: {str(e)}", exc_info=True)
                return False
            
    def _build_mylist_shop_rows_for_changes(self, added_list, updated_list):
        """
        추가 및 업데이트할 상가 행 데이터 생성 
        
        Args:
            added_list: 추가할 행 리스트
            updated_list: 업데이트할 행 리스트
            
        Returns:
            dict: 추가 및 업데이트할 행 데이터
        """
        if self.using_legacy:
            return self.legacy_logic._build_mylist_shop_rows_for_changes(added_list, updated_list)
        else:
            # 모듈식 구현에서는 독자적인 방식으로 처리
            try:
                self.logger.debug(f"모듈식 구현에서 상가 행 데이터 생성: 추가={len(added_list)}, 업데이트={len(updated_list)}")
                
                model = self.mylist_shop_model
                if not model:
                    self.logger.error("모델이 설정되지 않았습니다.")
                    return {"added": [], "updated": []}
                
                rowCount = model.rowCount()
                id_to_row = {}
                
                # ID와 해당 행 인덱스 매핑
                for r in range(rowCount):
                    item0 = model.item(r, 0)
                    if item0 and item0.data(Qt.UserRole+3) is not None:
                        id_to_row[item0.data(Qt.UserRole+3)] = r
                
                # 추가할 행 처리
                added_rows_parsed = []
                for add_obj in added_list:
                    temp_id = add_obj.get("temp_id")
                    row_idx = id_to_row.get(temp_id)
                    if row_idx is not None:
                        # 모듈식 구현에서 행 파싱 (데이터 모델 처리)
                        if hasattr(self.model, 'get_row_data'):
                            row_dict = self.model.get_row_data(row_idx)
                            if row_dict:
                                row_dict["temp_id"] = temp_id
                                added_rows_parsed.append(row_dict)
                        else:
                            # 폴백: 레거시 방식으로 파싱
                            self.logger.warning(f"모듈식 구현에서 get_row_data 메소드가 없어 레거시 방식 사용")
                            row_dict = self.legacy_logic._parse_mylist_shop_row(row_idx)
                            row_dict["temp_id"] = temp_id
                            added_rows_parsed.append(row_dict)
                
                # 업데이트할 행 처리
                updated_rows_parsed = []
                updated_ids = set(upd_obj.get("id") for upd_obj in updated_list 
                               if isinstance(upd_obj.get("id"), int) and upd_obj.get("id") > 0)
                
                for real_id in updated_ids:
                    row_idx = id_to_row.get(real_id)
                    if row_idx is not None:
                        # 모듈식 구현에서 행 파싱
                        if hasattr(self.model, 'get_row_data'):
                            row_dict = self.model.get_row_data(row_idx)
                            if row_dict:
                                row_dict["id"] = real_id
                                updated_rows_parsed.append(row_dict)
                        else:
                            # 폴백: 레거시 방식으로 파싱
                            self.logger.warning(f"모듈식 구현에서 get_row_data 메소드가 없어 레거시 방식 사용")
                            row_dict = self.legacy_logic._parse_mylist_shop_row(row_idx)
                            row_dict["id"] = real_id
                            updated_rows_parsed.append(row_dict)
                
                return {"added": added_rows_parsed, "updated": updated_rows_parsed}
                
            except Exception as e:
                self.logger.error(f"모듈식 구현에서 상가 행 데이터 생성 중 오류: {str(e)}", exc_info=True)
                # 오류 발생 시 레거시 방식으로 폴백
                self.logger.warning("레거시 방식으로 폴백하여 상가 행 데이터 생성")
                return self.legacy_logic._build_mylist_shop_rows_for_changes(added_list, updated_list)
            
    def on_data_loaded(self, rows):
        """
        데이터 로드 완료 처리
        
        Args:
            rows (list): 로드된 데이터 행
        """
        # 모델 데이터 설정
        self.model.append_rows(rows)
        if hasattr(self, 'autosave_status_label'):
            self.autosave_status_label.setText(f"데이터 로드 완료: {len(rows)}개 행")
        # 모드 전환 시 호환성 속성 업데이트
        self.mylist_shop_model = self.model.model
        
    def on_load_error(self, error_msg):
        """
        데이터 로드 오류 처리
        
        Args:
            error_msg (str): 오류 메시지
        """
        self.logger.error(f"데이터 로드 오류: {error_msg}")
        if hasattr(self, 'autosave_status_label'):
            self.autosave_status_label.setText(f"데이터 로드 오류: {error_msg}")
    
    def on_button_clicked(self, button_name):
        """
        버튼 클릭 이벤트 처리
        
        Args:
            button_name (str): 버튼 이름
        """
        if button_name == "add":
            # 행 추가 처리
            self.on_add_mylist_shop_row()
            self.logger.debug("add 버튼으로 행 추가 요청")
        elif button_name == "save":
            # 저장 처리를 container.save_handler로 위임
            if hasattr(self.container, 'save_handler'):
                self.container.save_handler.save_mylist_shop_changes()
                self.logger.debug("save 버튼으로 저장 요청")
            else:
                if hasattr(self, 'autosave_status_label'):
                    self.autosave_status_label.setText("저장 핸들러가 설정되지 않았습니다.")
        elif button_name == "refresh":
            # 데이터 다시 로드
            self.load_data()
            self.logger.debug("refresh 버튼으로 새로고침 요청")
        elif button_name == "naver_search":
            # 네이버 검수 기능 실행
            if hasattr(self.container, 'launch_naver_check_for_mylist'):
                self.container.launch_naver_check_for_mylist()
                self.logger.debug("네이버 검수 기능 요청")
            
    def filter_table_by_address(self, address_str):
        """
        주소로 테이블 필터링 (기존 MyListSangaLogic 호환)
        
        Args:
            address_str (str): 필터링할 주소
        """
        if self.using_legacy:
            return self.legacy_logic.filter_table_by_address(address_str)
        else:
            if hasattr(self.view_events, 'filter_by_address'):
                self.view_events.filter_by_address(address_str)
            else:
                self.logger.warning("filter_by_address 메서드를 찾을 수 없습니다.")
            
    # 추가 호환성 메서드들
    def toggle_mode(self):
        """
        레거시 모드와 모듈식 모드 전환 (테스트용)
        
        Returns:
            bool: 현재 모드 (True=레거시, False=모듈식)
        """
        self.using_legacy = not self.using_legacy
        
        # 모드 전환 시 호환성 속성 업데이트 - 직접 설정 불필요
        # 현재 모드에 맞는 값이 프로퍼티에서 자동으로 반환됨
        
        # 뷰 참조 업데이트
        if self.using_legacy:
            if hasattr(self.legacy_logic, 'mylist_shop_view'):
                self._mylist_shop_view = self.legacy_logic.mylist_shop_view
        else:
            # 모듈식 뷰 설정 유지
            pass
        
        self.logger.info(f"모드 전환: {'레거시' if self.using_legacy else '모듈식'}")
        return self.using_legacy
    
    def on_add_mylist_shop_row(self, initial_data=None):
        """
        상가 탭에 새 행 추가 (기존 MyListSangaLogic.on_add_mylist_shop_row와 호환)
        
        Args:
            initial_data: 초기 데이터 (선택 사항)
        """
        if self.using_legacy:
            return self.legacy_logic.on_add_mylist_shop_row(initial_data)
        else:
            try:
                # 모듈식 구현에서는 commands를 통해 처리
                self.commands.add_row(initial_data=initial_data)
                self.logger.debug("모듈식 구현으로 상가 행 추가 완료")
            except Exception as e:
                self.logger.error(f"모듈식 구현에서 상가 행 추가 중 오류: {str(e)}", exc_info=True)
                # 오류 발생 시 기존 방식으로 폴백
                return self.legacy_logic.on_add_mylist_shop_row(initial_data)
    
    def copy_mylist_shop_row(self, source_row_idx):
        """
        상가 행 복사 (기존 MyListSangaLogic.copy_mylist_shop_row와 호환)
        
        Args:
            source_row_idx: 복사할 행 인덱스
        """
        if self.using_legacy:
            return self.legacy_logic.copy_mylist_shop_row(source_row_idx)
        else:
            try:
                # 모듈식 구현에서는 소스 행의 데이터를 가져와서 새 행 추가
                model = self.mylist_shop_model
                if not model or source_row_idx < 0 or source_row_idx >= model.rowCount():
                    self.logger.warning(f"유효하지 않은 소스 행 인덱스: {source_row_idx}")
                    return
                
                # 소스 행 데이터 수집
                row_data = {}
                if hasattr(self.model, 'get_row_data'):
                    row_data = self.model.get_row_data(source_row_idx)
                else:
                    # 헤더 정보 수집
                    headers = []
                    for c in range(model.columnCount()):
                        header_item = model.horizontalHeaderItem(c)
                        if header_item:
                            headers.append(header_item.text())
                    
                    # 행 데이터 수집
                    for c in range(min(len(headers), model.columnCount())):
                        item = model.item(source_row_idx, c)
                        if item:
                            row_data[headers[c]] = item.text()
                
                # 새 행 추가 (초기 데이터와 함께)
                self.commands.add_row(initial_data=row_data)
                self.logger.debug(f"모듈식 구현으로 행 {source_row_idx} 복사 완료")
            except Exception as e:
                self.logger.error(f"모듈식 구현에서 상가 행 복사 중 오류: {str(e)}", exc_info=True)
                # 오류 발생 시 기존 방식으로 폴백
                return self.legacy_logic.copy_mylist_shop_row(source_row_idx)
                
    def update_mylist_shop_row_id(self, old_tid, new_id):
        """
        임시 ID를 실제 DB ID로 업데이트 (기존 MyListSangaLogic 호환)
        
        Args:
            old_tid: 기존 임시 ID
            new_id: 새 실제 ID
            
        Returns:
            bool: 성공 여부
        """
        if self.using_legacy:
            # 메서드 이름 수정 - "_update_mylist_shop_row_id" 대신 올바른 메서드 호출
            if hasattr(self.legacy_logic, 'update_mylist_shop_row_id'):
                # 직접 해당 메서드가 있으면 호출
                return self.legacy_logic.update_mylist_shop_row_id(old_tid, new_id)
            elif hasattr(self.legacy_logic, '_update_mylist_shop_row_id'):
                # 언더스코어가 있는 메서드가 있으면 호출 
                return self.legacy_logic._update_mylist_shop_row_id(old_tid, new_id)
            else:
                # 호환성 레이어에서 구현할 수 밖에 없는 경우
                self.logger.warning("레거시 로직에 ID 업데이트 메서드가 없습니다. 직접 구현을 시도합니다.")
                return self._update_id_directly(old_tid, new_id)
        else:
            try:
                # 모듈식 구현에서는 모델에서 직접 ID 업데이트
                model = self.mylist_shop_model
                if not model:
                    self.logger.error("모델이 설정되지 않았습니다.")
                    return False
                
                # 임시 ID로 행 찾기
                for r in range(model.rowCount()):
                    item0 = model.item(r, 0)
                    if not item0:
                        continue
                    
                    row_id = item0.data(Qt.UserRole + 3)
                    if row_id == old_tid:
                        # ID 업데이트
                        item0.setData(new_id, Qt.UserRole + 3)
                        self.logger.debug(f"모듈식 구현으로 ID 업데이트 완료: {old_tid} → {new_id}")
                        return True
                
                self.logger.warning(f"임시 ID {old_tid}를 가진 행을 찾을 수 없습니다.")
                return False
            except Exception as e:
                self.logger.error(f"모듈식 구현에서 ID 업데이트 중 오류: {str(e)}", exc_info=True)
                # 직접 구현 시도
                return self._update_id_directly(old_tid, new_id)
    
    def _update_id_directly(self, old_tid, new_id):
        """
        임시 ID를 실제 DB ID로 직접 업데이트하는 메서드
        (레거시/모듈식 메서드를 사용할 수 없을 때의 마지막 대안)
        
        Args:
            old_tid: 기존 임시 ID
            new_id: 새 실제 ID
            
        Returns:
            bool: 성공 여부
        """
        try:
            model = self.mylist_shop_model
            if not model:
                self.logger.error("직접 ID 업데이트 실패: 모델이 없습니다.")
                return False
            
            # 임시 ID로 행 찾기
            for r in range(model.rowCount()):
                item0 = model.item(r, 0)
                if not item0:
                    continue
                
                current_id = item0.data(Qt.UserRole + 3)
                if current_id == old_tid:
                    # ID 업데이트
                    item0.setData(new_id, Qt.UserRole + 3)
                    self.logger.debug(f"직접 구현으로 ID 업데이트 완료: {old_tid} → {new_id}")
                    return True
            
            self.logger.warning(f"직접 구현: 임시 ID {old_tid}를 가진 행을 찾을 수 없습니다.")
            return False
        except Exception as e:
            self.logger.error(f"직접 ID 업데이트 중 오류: {str(e)}", exc_info=True)
            return False
    
    def delete_selected_mylist_shop_rows(self):
        """선택된 상가 행 삭제 (기존 MyListSangaLogic 호환)"""
        if self.using_legacy:
            return self.legacy_logic.delete_selected_mylist_shop_rows()
        else:
            try:
                # 현재 선택된 행 인덱스 가져오기
                view = self.mylist_shop_view
                if not view:
                    self.logger.error("뷰가 설정되지 않았습니다.")
                    return
                
                sel_model = view.selectionModel()
                if not sel_model:
                    self.logger.error("선택 모델이 설정되지 않았습니다.")
                    return
                
                selected_indexes = sel_model.selectedIndexes()
                if not selected_indexes:
                    return
                
                involved_rows = set(idx.row() for idx in selected_indexes)
                if not involved_rows:
                    return
                
                # 삭제 확인 대화상자
                from PyQt5.QtWidgets import QMessageBox
                reply = QMessageBox.question(view, "삭제 확인", 
                    f"선택한 셀이 포함된 {len(involved_rows)}개 행 전체를 삭제 상태로 표시하시겠습니까?",
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply == QMessageBox.No:
                    return
                
                # 삭제 처리
                rows_to_mark = sorted(list(involved_rows))
                marked_count = 0
                
                for row_idx in rows_to_mark:
                    # 모듈식 구현에서는 commands.delete_rows를 사용
                    self.commands.delete_rows([row_idx])
                    marked_count += 1
                
                self.logger.info(f"모듈식 구현으로 {marked_count}개 상가 행 삭제 표시 완료")
                
            except Exception as e:
                self.logger.error(f"모듈식 구현에서 상가 행 삭제 중 오류: {str(e)}", exc_info=True)
                # 오류 발생 시 기존 방식으로 폴백
                return self.legacy_logic.delete_selected_mylist_shop_rows() 
    
    # 모듈식 모드에서 필요한 추가 메서드들
    def on_naver_search_clicked(self):
        """네이버 검색 버튼 클릭 처리"""
        if self.using_legacy:
            if hasattr(self.legacy_logic, 'on_naver_search_clicked'):
                return self.legacy_logic.on_naver_search_clicked()
        else:
            self.logger.info("네이버 검색 버튼 클릭됨")
            if hasattr(self.container, 'launch_naver_check_for_mylist'):
                self.container.launch_naver_check_for_mylist()
                
    def on_save_mylist_shop_changes(self):
        """저장 버튼 클릭 처리"""
        if self.using_legacy:
            if hasattr(self.legacy_logic, 'on_save_mylist_shop_changes'):
                return self.legacy_logic.on_save_mylist_shop_changes()
        else:
            self.logger.info("저장 버튼 클릭됨")
            if hasattr(self.container, 'save_handler'):
                self.container.save_handler.save_mylist_shop_changes()
                
    def export_selected_shop_to_excel(self):
        """엑셀 다운로드 버튼 클릭 처리"""
        if self.using_legacy:
            if hasattr(self.legacy_logic, 'export_selected_shop_to_excel'):
                return self.legacy_logic.export_selected_shop_to_excel()
        else:
            self.logger.info("엑셀 다운로드 버튼 클릭됨")
            view = self.mylist_shop_view
            model = self.mylist_shop_model
            
            if not view or not model:
                return
                
            # 저장 경로 대화상자 표시
            from PyQt5.QtWidgets import QFileDialog
            file_path, _ = QFileDialog.getSaveFileName(
                view, "엑셀 저장", "", "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return  # 사용자가 취소함
                
            try:
                # 엑셀 내보내기 로직
                import openpyxl
                from openpyxl import Workbook
                
                wb = Workbook()
                ws = wb.active
                ws.title = "상가 목록"
                
                # 헤더 추가
                headers = []
                for c in range(model.columnCount()):
                    header = model.horizontalHeaderItem(c)
                    if header:
                        headers.append(header.text())
                        
                for c, header in enumerate(headers, 1):
                    ws.cell(row=1, column=c, value=header)
                
                # 데이터 추가
                for r in range(model.rowCount()):
                    for c, header in enumerate(headers, 1):
                        item = model.item(r, c-1)
                        if item:
                            ws.cell(row=r+2, column=c, value=item.text())
                
                # 파일 저장
                wb.save(file_path)
                self.logger.info(f"엑셀 파일 저장 완료: {file_path}")
                
            except Exception as e:
                self.logger.error(f"엑셀 내보내기 중 오류: {str(e)}", exc_info=True)
                
    def on_open_sanga_tk_for_mylist_shop(self):
        """네이버부동산 검수 버튼 클릭 처리"""
        if self.using_legacy:
            if hasattr(self.legacy_logic, 'on_open_sanga_tk_for_mylist_shop'):
                return self.legacy_logic.on_open_sanga_tk_for_mylist_shop()
        else:
            self.logger.info("네이버부동산 검수 버튼 클릭됨")
            if hasattr(self.container, 'launch_naver_tk'):
                self.container.launch_naver_tk("sanga")
                
    def _get_horizontal_headers(self):
        """헤더 목록 반환"""
        if self.using_legacy:
            if hasattr(self.legacy_logic, '_get_horizontal_headers'):
                return self.legacy_logic._get_horizontal_headers()
        
        return ["주소", "담당자", "매물번호", "호실", "층", "입주가능일", "월세(만원)", 
                "보증금(만원)", "관리비(만원)", "권리금(만원)", "방/화장실", "평수(공급/전용)", 
                "현업종", "업종제한", "주차", "건물종류", "전용출입구", "전화번호", 
                "메모", "추가정보", "승인일자", "재광고"]
                
    def _mylist_shop_context_menu(self, pos):
        """
        상가 테이블 컨텍스트 메뉴 (기존 MyListSangaLogic._mylist_shop_context_menu 호환)
        
        Args:
            pos: 메뉴 위치 (QPoint)
        """
        if self.using_legacy:
            if hasattr(self.legacy_logic, '_mylist_shop_context_menu'):
                return self.legacy_logic._mylist_shop_context_menu(pos)
        else:
            # context_menu_events.show_context_menu 메서드 호출
            if hasattr(self.context_menu_events, 'show_context_menu'):
                return self.context_menu_events.show_context_menu(pos)
            else:
                self.logger.error("context_menu_events.show_context_menu 메서드를 찾을 수 없습니다.")
                
    def on_mylist_shop_current_changed(self, current, previous):
        """
        상가 테이블에서 선택된 행이 변경될 때 호출 (기존 MyListSangaLogic.on_mylist_shop_current_changed 호환)
        
        Args:
            current: 현재 선택된 인덱스
            previous: 이전에 선택된 인덱스
        """
        if self.using_legacy:
            if hasattr(self.legacy_logic, 'on_mylist_shop_current_changed'):
                self.legacy_logic.on_mylist_shop_current_changed(current, previous)
        else:
            # view_events.on_current_changed 메서드 호출
            if hasattr(self.view_events, 'on_current_changed'):
                self.view_events.on_current_changed(current, previous)
            else:
                self.logger.error("view_events.on_current_changed 메서드를 찾을 수 없습니다.")
        
        # 부모 컨테이너의 하단 테이블 업데이트 호출 (상가 행 클릭 시 하단 테이블 populate)
        try:
            if hasattr(self, 'container') and hasattr(self.container, 'update_selection_from_mylist'):
                print(f"[DEBUG] SangaBridge: 상가 행 선택됨 - 하단 테이블 업데이트 호출")
                self.container.update_selection_from_mylist("sanga")
            else:
                print(f"[DEBUG] SangaBridge: container 또는 update_selection_from_mylist 메서드 없음")
        except Exception as e:
            self.logger.error(f"하단 테이블 업데이트 호출 실패: {e}")
                
    def on_mylist_shop_view_double_clicked(self, index):
        """
        상가 테이블에서 아이템 더블 클릭 이벤트 처리 (기존 MyListSangaLogic.on_mylist_shop_view_double_clicked 호환)
        
        Args:
            index: 더블 클릭된 인덱스
        """
        if self.using_legacy:
            if hasattr(self.legacy_logic, 'on_mylist_shop_view_double_clicked'):
                return self.legacy_logic.on_mylist_shop_view_double_clicked(index)
        else:
            # view_events.on_view_double_clicked 메서드 호출
            if hasattr(self.view_events, 'on_view_double_clicked'):
                return self.view_events.on_view_double_clicked(index)
            else:
                self.logger.error("view_events.on_view_double_clicked 메서드를 찾을 수 없습니다.")
                
    def on_mylist_shop_header_clicked(self, logical_index):
        """
        상가 테이블 헤더 클릭 이벤트 처리 (기존 MyListSangaLogic.on_mylist_shop_header_clicked 호환)
        
        Args:
            logical_index: 클릭된 헤더 인덱스
        """
        if self.using_legacy:
            if hasattr(self.legacy_logic, 'on_mylist_shop_header_clicked'):
                return self.legacy_logic.on_mylist_shop_header_clicked(logical_index)
        else:
            # view_events.on_header_section_clicked 메서드 호출
            if hasattr(self.view_events, 'on_header_section_clicked'):
                return self.view_events.on_header_section_clicked(logical_index)
            else:
                self.logger.error("view_events.on_header_section_clicked 메서드를 찾을 수 없습니다.")
                
    def on_mylist_shop_change_status(self, pos):
        """
        상가 테이블에서 상태 변경 (기존 MyListSangaLogic.on_mylist_shop_change_status 호환)
        
        Args:
            pos: 메뉴 위치
        """
        if self.using_legacy:
            if hasattr(self.legacy_logic, 'on_mylist_shop_change_status'):
                return self.legacy_logic.on_mylist_shop_change_status(pos)
        else:
            # context_menu_events.change_status 메서드 호출
            if hasattr(self.context_menu_events, 'change_status'):
                return self.context_menu_events.change_status(pos)
            else:
                self.logger.error("context_menu_events.change_status 메서드를 찾을 수 없습니다.")
    
    def populate_mylist_shop_table(self, rows_data):
        """
        배치 데이터로 상가 테이블 populate
        
        Args:
            rows_data: 배치 API에서 받은 상가 데이터 리스트
        """
        try:
            self.logger.info(f"[SangaBridge] populate_mylist_shop_table 호출됨: {len(rows_data)}개 행")
            
            if self.using_legacy:
                # 레거시 로직 사용
                if hasattr(self.legacy_logic, 'mylist_shop_dict'):
                    # 기존 dict에 데이터 저장
                    for row in rows_data:
                        address = f"{row.get('dong', '')} {row.get('jibun', '')}".strip()
                        if address not in self.legacy_logic.mylist_shop_dict:
                            self.legacy_logic.mylist_shop_dict[address] = []
                        self.legacy_logic.mylist_shop_dict[address].append(row)
                    
                    # 테이블 업데이트
                    if hasattr(self.legacy_logic, 'filter_and_populate'):
                        self.legacy_logic.filter_and_populate()
                        self.logger.info("[SangaBridge] 레거시 로직으로 테이블 populate 완료")
                else:
                    self.logger.warning("[SangaBridge] 레거시 로직에 mylist_shop_dict가 없음")
            else:
                # 모듈식 구현 사용
                self.model.clear_model()
                self.model.append_rows(rows_data)
                self.logger.info("[SangaBridge] 모듈식 로직으로 테이블 populate 완료")
                
        except Exception as e:
            self.logger.error(f"[SangaBridge] populate_mylist_shop_table 실패: {e}", exc_info=True) 