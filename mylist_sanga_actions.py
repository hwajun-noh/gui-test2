# mylist_sanga_actions.py
import os
import sys
import glob
import time
import json
from datetime import datetime
import openpyxl  # 엑셀 다운로드/업로드 시 사용
from openpyxl import Workbook

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt, QModelIndex, QUrl, pyqtSlot
from PyQt5.QtGui import QStandardItem
from PyQt5.QtWidgets import QMessageBox, QInputDialog, QFileDialog, QApplication

from dialogs import NaverShopSearchDialog
import traceback
import requests # requests 임포트 추가
import logging # logging 임포트 추가

def on_save_mylist_shop_changes(logic_instance):
    """Forwards the save request to the container's save_handler."""
    if logic_instance.container and logic_instance.container.save_handler:
        logic_instance.logger.info("Manual save requested for shop tab. Forwarding to save_handler.")
        logic_instance.container.save_handler.save_pending_shop_changes()
    else:
         logic_instance.logger.error("Cannot save shop changes: container or save_handler is missing.")
         QMessageBox.warning(logic_instance.parent_app, "오류", "저장 핸들러를 찾을 수 없어 저장할 수 없습니다.")

def on_naver_search_clicked(logic_instance):
    """Opens the Naver property search dialog."""
    logic_instance.logger.info("MyListSanga: 'Naver Property Search' button clicked.")

    # --- 다이얼로그 재사용 또는 새로 생성 ---
    try:
        # 객체가 삭제되었지만 참조가 남아있는 경우 처리
        if hasattr(logic_instance, 'naver_search_dialog'):
            dialog_ref = logic_instance.naver_search_dialog
            
            # 안전하게 객체 존재 여부 및 가시성 확인
            is_visible = False
            try:
                is_visible = dialog_ref is not None and dialog_ref.isVisible()
            except RuntimeError:
                # C++ 객체가 이미 삭제된 경우 - 참조를 제거합니다
                logic_instance.logger.warning("Detected deleted dialog reference, will create a new one.")
                logic_instance.naver_search_dialog = None
                
            if is_visible:
                logic_instance.logger.info("Existing Naver Search Dialog found for MyListSanga. Activating.")
                dialog_ref.activateWindow()
                dialog_ref.raise_()
                return # 이미 열려있으면 새로 생성하지 않음
    except Exception as e:
        logic_instance.logger.error(f"Error checking existing dialog: {e}", exc_info=True)
        # 오류 발생 시 이전 참조를 정리하고 새로 생성
        logic_instance.naver_search_dialog = None
    # ------------------------------------

    # --- 다이얼로그 인스턴스를 logic_instance 속성으로 저장 ---
    logic_instance.naver_search_dialog = NaverShopSearchDialog(
        parent_app=logic_instance.parent_app,
        mylist_tab=logic_instance.container 
    )
    
    # 대화상자가 삭제될 때 참조를 정리하기 위한 시그널 연결
    logic_instance.naver_search_dialog.destroyed.connect(
        lambda: setattr(logic_instance, 'naver_search_dialog', None)
    )
    # --------------------------------------------------

    # <<< 시그널 연결 추가 >>>
    if hasattr(logic_instance.parent_app, 'all_tab') and hasattr(logic_instance.parent_app.all_tab, 'search_by_address'):
        # --- 수정: 메인 앱의 중앙 핸들러에 연결 ---
        if hasattr(logic_instance.parent_app, 'handle_address_selection_from_dialog'):
            logic_instance.naver_search_dialog.addressClicked.connect(logic_instance.parent_app.handle_address_selection_from_dialog)
            logic_instance.logger.info("Connected NaverShopSearchDialog.addressClicked to main_app.handle_address_selection_from_dialog")
        else:
            logic_instance.logger.warning("Could not connect addressClicked signal: main_app.handle_address_selection_from_dialog slot not found.")
        # logic_instance.naver_search_dialog.addressClicked.connect(logic_instance.parent_app.all_tab.search_by_address) # 이전 연결 주석 처리
        # logic_instance.logger.info("Connected NaverShopSearchDialog.addressClicked to AllTab.search_by_address") # 이전 로그 주석 처리
        # ------------------------------------
    else:
        logic_instance.logger.warning("Could not connect addressClicked signal: all_tab or search_by_address slot not found.")
    # <<< --- >>>

    logic_instance.naver_search_dialog.show()
    logic_instance.logger.info("Naver Search Dialog shown (modeless).")
    
    # if result == QMessageBox.Accepted:
    #     # User clicked OK, get the result and add to the model

def add_shop_row_with_data(logic_instance, row_dict_from_naver):
    """Creates a new shop row filled with data from Naver search."""
    if not row_dict_from_naver:
        return
    
    # Call container method to add the row, specifying Naver format parsing
    if logic_instance.container:
        logic_instance.container.add_new_shop_row(initial_data=row_dict_from_naver, parse_naver_format=True)

def filter_table_by_address(logic_instance, address_str: str):
    """
    Filters the shop table to only show rows containing the given address string.
    """
    if not address_str:
        return  # Nothing to filter
    
    address_str = address_str.strip().lower()  # Case-insensitive
    
    model = logic_instance.mylist_shop_model
    view = logic_instance.mylist_shop_view
    
    if not model or not view:
        return
    
    # Determine column index for the address column
    headers = [model.horizontalHeaderItem(c).text() for c in range(model.columnCount())]
    try:
        addr_col_idx = headers.index("주소")
    except ValueError:
        print("[WARN] Cannot filter by address: '주소' column not found")
        return
    
    # This is a simple filter that just checks if the address_str is in the cell text
    # Could be improved with more sophisticated matching
    for row in range(model.rowCount()):
        item = model.item(row, addr_col_idx)
        if not item:
            continue
        
        cell_text = item.text().lower()
        row_visible = address_str in cell_text
        
        # Hide/show row based on filter match
        view.setRowHidden(row, not row_visible)
    
    matching_count = sum(1 for row in range(model.rowCount()) if not view.isRowHidden(row))
    print(f"[INFO] Address filter applied: '{address_str}' - {matching_count} rows visible")

def copy_mylist_shop_row(logic_instance, source_row_idx):
    """Copies a row and adds it as a new temporary row."""
    model = logic_instance.mylist_shop_model
    if not model or source_row_idx < 0 or source_row_idx >= model.rowCount():
        return

    col_count = model.columnCount()
    copied_values = [model.item(source_row_idx, c).text() if model.item(source_row_idx, c) else "" for c in range(col_count)]

    # Use container method to add row and handle pending state & temp ID
    if logic_instance.container:
        logic_instance.container.add_new_shop_row(initial_data=copied_values)

def export_selected_shop_to_excel(logic_instance):
    """
    Exports selected rows to an Excel file based on predefined mapping.
    Uses logic similar to the previous version:
    - Includes specific columns defined in excel_map.
    - Applies specific data transformations (무권리, 공실, 사용승인일, 담당자).
    - Saves file automatically with a generated name.
    - Uses print for feedback.
    """
    view = logic_instance.mylist_shop_view
    model = logic_instance.mylist_shop_model

    if not view or not model:
        print("View 또는 Model을 찾을 수 없습니다.")
        return

    # 1) selection: 단일 셀만 클릭해도 해당 row 포함
    selection_model = view.selectionModel()
    if not selection_model:
        print("Selection Model을 찾을 수 없습니다.")
        return

    selected_indexes = selection_model.selectedIndexes()
    if not selected_indexes:
        print("선택된 셀이 없습니다.")
        return

    # 선택된 셀들의 row 인덱스 -> 중복 제거 -> 오름차순 정렬
    row_set = set(idx.row() for idx in selected_indexes)
    row_indexes = sorted(row_set)
    if not row_indexes:
        print("선택된 행이 없습니다.")
        return

    # (A) 엑셀 열 매핑: (엑셀열 인덱스, "엑셀헤더명", mylist col 인덱스)
    excel_map = [
        (0, "주소", 0),
        (1, "호", 1),
        (2, "권리금", 5),
        (3, "현업종", 6),
        (4, "보증금/월세", 3),
        (5, "관리비(만원)", 4),
        (6, "평수(㎡)", 7),
        (7, "해당층/총층", 2),
        (8, "담당자", 10),
        (9, "연락처", 8),
        (10, "방/화장실", 15),
        (11, "주차대수", 12),
        (12, "매물번호", 9),
        (13, "용도", 13),
        (14, "사용승인일", 14), # 하이픈 제거 필요
        (15, "소유자명", 18),
        (16, "관계", 19),
        (17, "사진경로", 17),
        (18, "광고종료일", 16),
        (19, "메모", 11), # 이전 코드처럼 메모는 포함 (이전 코드 주석은 잘못되었음)
    ]

    # '재광고' 상태를 가져올 모델 컬럼 인덱스 확인 필요
    # 이전 코드에서는 20번 인덱스를 사용했으나, 현재 모델 구조 확인 필요
    # 우선 현재 테이블의 '재광고' 헤더를 찾아 인덱스를 동적으로 결정
    all_headers = [model.horizontalHeaderItem(c).text() for c in range(model.columnCount())]
    try:
        re_ad_col_index = all_headers.index("재광고")
    except ValueError:
        print("[WARN] '재광고' 컬럼을 찾을 수 없어 담당자명 처리가 제한될 수 있습니다.")
        re_ad_col_index = -1 # 재광고 컬럼 못 찾음

    # (B) Workbook 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "마이리스트_상가"

    # (C) 엑셀 첫 행(1행) => 헤더 작성 (excel_map 기준)
    for excel_col_idx, header_text, _ in excel_map:
        ws.cell(row=1, column=excel_col_idx + 1).value = header_text

    # (D) 엑셀 row=2부터 실제 mylist_shop 데이터 (excel_map 기준)
    start_excel_row = 2
    for i, row_idx in enumerate(row_indexes):
        excel_row = start_excel_row + i

        # 재광고 상태 값 가져오기 (담당자 처리용)
        re_ad_val = ""
        if re_ad_col_index != -1:
            re_ad_item = model.item(row_idx, re_ad_col_index)
            # 현재 '재광고' 컬럼은 Y/N 대신 "재광고" 문자열 또는 빈 값일 수 있음
            re_ad_val = re_ad_item.text().strip() if re_ad_item else ""

        # excel_map 순서대로 값 가져와서 변환 후 대입
        for excel_col_idx, _, mylist_col_idx in excel_map:
            item_ = model.item(row_idx, mylist_col_idx)
            text_val = item_.text() if item_ else ""

            # === 이전 코드 데이터 변환 로직 적용 ===
            # 권리금(col=5): "무권리" -> ""
            if mylist_col_idx == 5 and text_val.strip() == "무권리":
                text_val = ""
            # 현업종(col=6): "공실" -> ""
            elif mylist_col_idx == 6 and text_val.strip() == "공실":
                text_val = ""
            # 사용승인일(col=14): 하이픈 제거
            elif mylist_col_idx == 14:
                text_val = text_val.replace("-", "")
            # 담당자(col=10): 재광고 상태에 따라 형식 변경
            elif mylist_col_idx == 10:
                manager_name = text_val.strip()
                if manager_name:
                    # first_char = manager_name[0] # 첫 글자 추출 제거
                    # 재광고 상태 확인 (이전 코드 로직과 유사하게)
                    # 현재 '재광고' 컬럼 값이 "재광고" 문자열이면 재광고로 판단
                    if re_ad_val == "재광고":
                        # text_val = f"{first_char}(재광고)" # 재광고 표시 로직 제거
                        text_val = f"{manager_name}(재광고)" # 전체 이름 뒤에 (재광고) 추가
                    else:
                        # text_val = first_char # 첫 글자만 사용하는 로직 제거
                        text_val = manager_name # 전체 이름 사용
                else:
                    text_val = "" # 담당자 비었으면 ""
            # ======================================

            ws.cell(row=excel_row, column=excel_col_idx + 1).value = text_val

    # (E) 파일명 자동 생성 및 저장 (이전 코드 방식)
    try:
        # logic_instance에서 현재 매니저 이름 가져오기 (가정)
        # 실제 속성 이름 확인 필요 (예: logic_instance.current_manager)
        current_manager_name = getattr(logic_instance, 'current_manager', 'UnknownManager')
        now_str = datetime.now().strftime("%Y_%m_%d_%H%M")
        manager_str = current_manager_name.replace(" ", "_")
        filename = f"{now_str}_{manager_str}.xlsx"

        wb.save(filename)
        print(f"[INFO] 엑셀 다운로드 완료 ({len(row_indexes)} 행) => {filename}")

    except AttributeError:
         print("[ERROR] 현재 매니저 이름 ('current_manager')을 가져올 수 없어 파일명 생성이 제한됩니다.")
         # 파일 저장 시도 (기본 이름 사용)
         try:
             fallback_filename = f"마이리스트_상가_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
             wb.save(fallback_filename)
             print(f"[INFO] 엑셀 다운로드 완료 ({len(row_indexes)} 행) => {fallback_filename}")
         except Exception as e_save:
             print(f"[ERROR] 엑셀 파일 저장 중 오류 발생: {e_save}")
    except Exception as e:
        print(f"[ERROR] 엑셀 파일 저장 중 오류 발생: {e}")

def on_open_sanga_tk_for_mylist_shop(logic_instance):
    """
    Opens the TKinter window for Naver property inspection.
    
    새로운 방식: 컨테이너의 launch_naver_check_for_mylist 함수를 호출하여
    모든 행 데이터를 전달하여 다음 버튼으로 전체 행을 탐색할 수 있게 합니다.
    """
    try:
        # 현재 선택된 행 확인
        view = logic_instance.mylist_shop_view
        if not view or view.currentIndex().row() < 0:
            QMessageBox.information(
                logic_instance.parent_app,
                "선택 없음",
                "검수할 행을 선택하세요."
            )
            return
        
        # 컨테이너의 네이버부동산 검수 실행 함수 호출
        if logic_instance.container and hasattr(logic_instance.container, 'launch_naver_check_for_mylist'):
            logic_instance.container.launch_naver_check_for_mylist()
        else:
            # 컨테이너가 없거나 메서드가 없는 경우
            print("[ERROR] MyListContainer 또는 launch_naver_check_for_mylist 메서드를 찾을 수 없습니다.")
            QMessageBox.critical(
                logic_instance.parent_app,
                "실행 오류",
                "네이버부동산 검수 기능을 실행할 수 없습니다. MyListContainer 구성을 확인하세요."
            )
            
    except Exception as e:
        print(f"[ERROR] 네이버부동산 검수 실행 중 오류: {e}")
        import traceback
        print(traceback.format_exc())
        
        QMessageBox.critical(
            logic_instance.parent_app,
            "검수 창 실행 오류",
            f"네이버부동산 검수 창을 실행하는 중 오류가 발생했습니다:\n{e}"
        )

def highlight_mylist_shop_row_by_id(logic_instance, pk_id):
    """
    Finds and highlights a row by ID.
    This should be a QtCore.pyqtSlot(int) in the original class.
    """
    row_index = logic_instance.find_mylist_shop_row_by_id(pk_id)
    view = logic_instance.mylist_shop_view
    
    if row_index is not None and view:
        view.selectRow(row_index)
        view.scrollTo(view.model().index(row_index, 0))
        print(f"[INFO] Highlighted row {row_index} with ID {pk_id}")
    else:
        print(f"[WARN] Could not find row with ID {pk_id} to highlight")

def _handle_save_complete(logic_instance, result):
    """(Main Thread Slot) Handles the completion of the save operation."""
    # print(f"[DEBUG] _handle_save_complete called with result: {result}")
    logic_instance.logger.info(f"_handle_save_complete: Received save result (status={result.get('status') if result else 'None'}).")

    if not result:
        logic_instance.logger.error("_handle_save_complete: Received empty result.")
        QMessageBox.critical(logic_instance.parent_app, "저장 실패", "서버로부터 응답이 없습니다.")
        logic_instance.container.update_autosave_status("저장 실패 (응답 없음)")
        return

    if result.get("status") == "ok":
        logic_instance.logger.info("_handle_save_complete: Save successful. Processing updates.")
        added_results = result.get("added_ids", {})  # {temp_id: new_db_id}
        updated_count = result.get("updated_count", 0)
        deleted_count = result.get("deleted_count", 0)

        # Process added items: Update temporary IDs to real DB IDs
        model = logic_instance.mylist_shop_model
        if model:
            for temp_id_str, new_id in added_results.items():
                try:
                    temp_id = int(temp_id_str)
                    logic_instance.logger.debug(f"Updating temp_id {temp_id} to new_id {new_id}")
                    logic_instance._update_mylist_shop_row_id(temp_id, new_id)
                    # 새로 추가된 행의 배경색도 초기화 (기본 배경색일 가능성이 높지만 안전하게 처리)
                    row_idx = logic_instance.find_mylist_shop_row_by_id(new_id)
                    if row_idx is not None:
                        for col in range(model.columnCount()):
                            item = model.item(row_idx, col)
                            if item:
                                item.setBackground(QtGui.QBrush(Qt.NoBrush)) # 배경색 초기화
                except ValueError:
                     logic_instance.logger.error(f"_handle_save_complete: Invalid temp_id format '{temp_id_str}' received.")
                except Exception as e_update:
                     logic_instance.logger.error(f"_handle_save_complete: Error updating row ID for temp_id {temp_id_str}: {e_update}", exc_info=True)

            # 저장 성공한 업데이트 항목 배경색 복원
            # 저장 *전*의 pending 상태를 가져와야 함
            pending_updates_before_save = logic_instance.container.pending_manager.get_pending_shop_changes().get("updated", [])
            for update_item in pending_updates_before_save:
                record_id = update_item.get("id")
                if record_id is None: continue # ID 없으면 스킵

                row_idx = logic_instance.find_mylist_shop_row_by_id(record_id)
                if row_idx is not None:
                    # update_item 딕셔너리에 포함된 모든 필드에 대해 배경색 복원 시도
                    for db_field_name in update_item.keys():
                        if db_field_name == "id": continue # ID 필드는 스킵

                        display_col_name = None
                        # DB 필드명 -> 표시 열 이름 변환 (주의: 복합 필드는 이 로직으로 처리 안됨)
                        for disp_name, db_name in logic_instance.parent_app.COLUMN_MAP_MYLIST_SHOP_DISPLAY_TO_DB.items():
                            if db_name == db_field_name:
                                display_col_name = disp_name
                                break
                        
                        # 복합 필드 (예: 보증금/월세) 처리 로직 추가 필요 (여기서는 생략)
                        # if db_field_name in ["deposit", "rent"]: display_col_name = "보증금/월세"
                        # elif ...

                        if display_col_name:
                            try:
                                headers = [model.horizontalHeaderItem(c).text() for c in range(model.columnCount())]
                                col_idx = headers.index(display_col_name)
                                item = model.item(row_idx, col_idx)
                                if item and item.background() != QtGui.QBrush(Qt.NoBrush): # 이미 초기화된 경우는 제외
                                    item.setBackground(QtGui.QBrush(Qt.NoBrush)) # 배경색 초기화
                                    logic_instance.logger.debug(f"Reset background for updated item: ID={record_id}, Row={row_idx}, Col={col_idx} ({display_col_name})")
                            except ValueError:
                                logic_instance.logger.warning(f"Could not find column index for display name '{display_col_name}' (DB field: {db_field_name}) while resetting background.")
                            except Exception as e_reset:
                                logic_instance.logger.error(f"Error resetting background for item (ID:{record_id}, Row:{row_idx}, Col:{display_col_name}): {e_reset}")
                        # else:
                        #     logic_instance.logger.warning(f"Could not map DB field '{db_field_name}' to display column name for background reset.")


        # Clear the pending state in the manager for successfully saved items
        # 현재는 성공 시 모든 pending 상태를 클리어함 (개별 클리어 로직은 복잡하여 추후 개선 필요)
        logic_instance.container.pending_manager.clear_shop_pending_state()

        status_msg = f"저장 완료 (추가: {len(added_results)}, 수정: {updated_count}, 삭제: {deleted_count})"
        QMessageBox.information(logic_instance.parent_app, "저장 완료", status_msg)
        logic_instance.container.update_autosave_status(status_msg)

        # Reload data after successful save? Maybe optional?
        # logic_instance.load_data()

    else:
        error_msg = result.get("message", "알 수 없는 서버 오류")
        logic_instance.logger.error(f"_handle_save_complete: Save failed. Server message: {error_msg}")
        QMessageBox.critical(logic_instance.parent_app, "저장 실패", f"서버 오류:\n{error_msg}")
        logic_instance.container.update_autosave_status(f"저장 실패 ({error_msg[:30]}...)") 

def run_naver_inspect_app(AppClass, params):
    """내부 함수: NaverSangaInspectApp 실행 및 예외 처리"""
    try:
        # SangaCheckAppMylist.run()은 data_list를 반환
        # 해당 결과로 추가 처리 가능
        result_data = AppClass(**params).run()
        print(f"[INFO] 검수 작업 완료: {len(result_data)}개 항목 처리됨")
        return result_data
    except Exception as run_e:
        print(f"[ERROR] 검수 창 실행 오류: {run_e}")
        import traceback
        print(traceback.format_exc())
        return None 