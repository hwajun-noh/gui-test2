#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
GUI 엑셀 파일 선택기 - 초고속 + 미리보기 버전
최소한의 기능으로 극한 성능 최적화 + 필수 미리보기 추가
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
import logging
from typing import Optional, Tuple
import os
import json
import openpyxl
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# 설정 파일 경로
CONFIG_FILE = "gui_config.json"


class FastExcelSelector:
    """초고속 엑셀 파일 선택기 + 미리보기"""
    
    def __init__(self):
        self.root = tk.Tk()
        self.selected_file = None
        self.selected_sheet = None
        self.start_row = 3  # 기본 시작 행
        self.config = self._load_config()
        
        self._setup_ui()
        
    def _load_config(self) -> dict:
        """설정 파일 로드"""
        default_config = {
            "last_directory": os.path.expanduser("~/Documents"),
            "window_geometry": "800x1100",  # 세로가 더 긴 창으로 변경
            "start_row": 3
        }
        
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                
                # 가로가 더 긴 설정이면 세로가 더 긴 설정으로 강제 업데이트
                geometry = config.get("window_geometry", "800x1100")
                width, height = map(int, geometry.split('x'))
                if width >= height:  # 가로가 세로보다 크거나 같으면
                    config["window_geometry"] = "800x1100"
                    print("🔄 GUI 창 크기를 세로가 더 긴 형태(800x1100)로 업데이트했습니다.")
                
                return config
        except:
            pass
        
        return default_config
    
    def _save_config(self):
        """설정 파일 저장"""
        try:
            self.config["start_row"] = self.start_row
            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, ensure_ascii=False, indent=2)
        except:
            pass
        
    def _setup_ui(self):
        """최적화된 UI 구성 + 미리보기"""
        self.root.title("엑셀 파일 선택 및 설정")
        
        # 창 크기 (미리보기 포함)
        geometry = self.config.get("window_geometry", "900x1200")
        self.root.geometry(geometry)
        self.root.resizable(True, True)
        
        # 화면 상단 중앙 배치 및 최소화 방지
        self.root.update_idletasks()
        screen_width = self.root.winfo_screenwidth()
        
        # 창 크기에서 너비 추출
        width = int(geometry.split('x')[0])
        x = (screen_width - width) // 2
        y = 50  # 화면 상단에서 50px 아래
        self.root.geometry(f"{geometry}+{x}+{y}")
        
        # 창을 맨 앞으로 가져오기 및 포커스 설정
        self.root.lift()
        self.root.attributes('-topmost', True)
        self.root.after(100, lambda: self.root.attributes('-topmost', False))
        self.root.focus_force()
        
        # 메인 프레임
        main_frame = tk.Frame(self.root, padx=15, pady=15)
        main_frame.pack(fill="both", expand=True)
        
        # 1. 파일 선택
        tk.Label(main_frame, text="1. 엑셀 파일 선택", font=("맑은 고딕", 10, "bold")).pack(anchor="w", pady=(0, 5))
        
        file_frame = tk.Frame(main_frame)
        file_frame.pack(fill="x", pady=(0, 10))
        
        self.file_label = tk.Label(file_frame, text="선택된 파일: 없음", fg="gray")
        self.file_label.pack(anchor="w", pady=(0, 5))
        
        tk.Button(file_frame, text="파일 선택", command=self._select_file, 
                 bg="#0078d4", fg="white", relief="flat", padx=20).pack(anchor="w")
        
        # 2. 시트 선택
        tk.Label(main_frame, text="2. 시트 선택", font=("맑은 고딕", 10, "bold")).pack(anchor="w", pady=(10, 5))
        
        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(main_frame, textvariable=self.sheet_var, state="disabled", width=80)
        self.sheet_combo.pack(anchor="w", pady=(0, 10))
        self.sheet_combo.bind('<<ComboboxSelected>>', self._on_sheet_selected)
        
        # 3. 시작 행 설정
        tk.Label(main_frame, text="3. 데이터 시작 행 설정", font=("맑은 고딕", 10, "bold")).pack(anchor="w", pady=(10, 5))
        
        start_row_frame = tk.Frame(main_frame)
        start_row_frame.pack(fill="x", pady=(0, 10))
        
        tk.Label(start_row_frame, text="데이터가 시작되는 행 번호:").pack(side="left")
        
        self.start_row_var = tk.IntVar(value=self.config.get("start_row", 3))
        self.start_row_spinbox = tk.Spinbox(start_row_frame, from_=2, to=20, width=5, 
                                           textvariable=self.start_row_var, 
                                           command=self._on_start_row_changed)
        self.start_row_spinbox.pack(side="left", padx=(5, 10))
        
        tk.Label(start_row_frame, text="(헤더는 1행, 데이터는 보통 2-3행부터)", fg="gray").pack(side="left")
        
        # 4. 빠른 미리보기
        tk.Label(main_frame, text="4. 데이터 미리보기 (상위 10행)", font=("맑은 고딕", 10, "bold")).pack(anchor="w", pady=(10, 5))
        
        preview_frame = tk.Frame(main_frame)
        preview_frame.pack(fill="both", expand=True, pady=(0, 10))
        
        # 미리보기 TreeView (최적화)
        self.preview_tree = ttk.Treeview(preview_frame, height=8, show='headings')
        self.preview_tree.pack(side="left", fill="both", expand=True)
        
        # 스크롤바 (세로만)
        scrollbar_y = ttk.Scrollbar(preview_frame, orient="vertical", command=self.preview_tree.yview)
        scrollbar_y.pack(side="right", fill="y")
        self.preview_tree.configure(yscrollcommand=scrollbar_y.set)
        
        # 5. 파일 정보 및 헤더 검증
        tk.Label(main_frame, text="5. 파일 정보 및 헤더 검증", font=("맑은 고딕", 10, "bold")).pack(anchor="w", pady=(10, 5))
        
        self.info_text = tk.Text(main_frame, height=4, width=80, bg="#f8f9fa", relief="flat", 
                                font=("맑은 고딕", 9))
        self.info_text.pack(fill="x", pady=(0, 10))
        self.info_text.insert("1.0", "파일을 선택하면 정보가 표시됩니다.")
        self.info_text.config(state="disabled")
        
        # 6. 상태 및 버튼
        status_frame = tk.Frame(main_frame)
        status_frame.pack(fill="x", pady=(10, 0))
        
        self.status_label = tk.Label(status_frame, text="파일을 선택해주세요.", fg="blue")
        self.status_label.pack(anchor="w", pady=(0, 10))
        
        button_frame = tk.Frame(status_frame)
        button_frame.pack(anchor="w")
        
        tk.Button(button_frame, text="취소", command=self._cancel, 
                 relief="flat", padx=20).pack(side="left", padx=(0, 10))
        
        self.ok_button = tk.Button(button_frame, text="확인", command=self._confirm, 
                                  bg="#28a745", fg="white", relief="flat", padx=20, state="disabled")
        self.ok_button.pack(side="left")
        
        # 창 닫기 이벤트
        self.root.protocol("WM_DELETE_WINDOW", self._on_closing)
        
        # 초기 미리보기 메시지
        self._show_preview_message("파일을 선택하면 미리보기가 표시됩니다.")
        
    def _on_start_row_changed(self):
        """시작 행 변경 이벤트"""
        self.start_row = self.start_row_var.get()
        # 시작 행 변경 시에는 미리보기 재로딩 불필요
        
    def _show_preview_message(self, message):
        """미리보기 영역에 메시지 표시"""
        # TreeView 초기화
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        
        self.preview_tree['columns'] = ('message',)
        self.preview_tree.heading('message', text='미리보기')
        self.preview_tree.column('message', width=800, anchor='center')
        self.preview_tree.insert('', 'end', values=(message,))
        
    def _select_file(self):
        """초고속 파일 선택"""
        try:
            # 최근 디렉토리 사용
            initial_folder = self.config.get("last_directory", os.path.expanduser("~/Documents"))
            
            file_path = filedialog.askopenfilename(
                parent=self.root,
                title="엑셀 파일 선택",
                initialdir=initial_folder,
                filetypes=[("Excel", "*.xlsx *.xls"), ("모든 파일", "*.*")]
            )
            
            if file_path:
                self.selected_file = file_path
                self.file_label.config(text=f"선택된 파일: {Path(file_path).name}", fg="green")
                
                # 디렉토리 저장
                self.config["last_directory"] = str(Path(file_path).parent)
                self._save_config()
                
                # 즉시 시트 로드
                self._load_sheets_fast()
                
        except Exception as e:
            messagebox.showerror("오류", f"파일 선택 오류: {e}")
    
    def _load_sheets_fast(self):
        """초고속 시트 로드"""
        try:
            self.status_label.config(text="시트 로딩 중...", fg="orange")
            self._show_preview_message("시트 정보를 불러오는 중...")
            self.root.update()  # 즉시 UI 업데이트
            
            # openpyxl로 빠르게 시트명만 가져오기
            wb = load_workbook(self.selected_file, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            
            # 콤보박스 업데이트
            self.sheet_combo['values'] = sheet_names
            self.sheet_combo['state'] = "readonly"
            
            if len(sheet_names) == 1:
                # 시트 1개면 자동 선택
                self.sheet_combo.current(0)
                self.selected_sheet = sheet_names[0]
                self._load_preview_fast()
            else:
                self.status_label.config(text=f"{len(sheet_names)}개 시트 - 하나를 선택하세요", fg="orange")
                self._show_preview_message("시트를 선택하면 미리보기가 표시됩니다.")
                
        except Exception as e:
            self.status_label.config(text="파일 읽기 실패", fg="red")
            self._show_preview_message("파일을 읽을 수 없습니다.")
            messagebox.showerror("오류", f"시트 로드 실패: {e}")
    
    def _on_sheet_selected(self, event):
        """시트 선택 이벤트"""
        try:
            self.selected_sheet = self.sheet_var.get()
            if self.selected_sheet:
                self._load_preview_fast()
        except Exception as e:
            messagebox.showerror("오류", f"시트 선택 오류: {e}")
    
    def _load_preview_fast(self):
        """초고속 미리보기 로드 (openpyxl 사용) - 10행까지"""
        try:
            if not self.selected_file or not self.selected_sheet:
                return
                
            self.status_label.config(text="미리보기 로딩 중...", fg="orange")
            self._show_preview_message("데이터를 불러오는 중...")
            self.root.update()
            
            # openpyxl로 빠르게 상위 11행까지 읽기 (헤더 + 10행)
            wb = load_workbook(self.selected_file, read_only=True)
            ws = wb[self.selected_sheet]
            
            # 헤더 (첫 번째 행) 가져오기 - 더 많은 컬럼 읽기
            headers = []
            for cell in ws[1]:
                if cell.value is not None:
                    headers.append(str(cell.value))
                else:
                    headers.append("")
                if len(headers) >= 30:  # 최대 30개 컬럼 (Z열까지 + 여유분)
                    break
            
            # 끝부분의 빈 헤더들 제거
            while headers and headers[-1] == "":
                headers.pop()
            
            # 데이터 (2-11행) 가져오기 - 10행으로 확장
            data_rows = []
            for row_num in range(2, 12):  # 2-11행 (10행)
                try:
                    row_data = []
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        if cell.value is not None:
                            # 텍스트 길이 제한
                            text = str(cell.value)[:30]
                            row_data.append(text)
                        else:
                            row_data.append("")
                    data_rows.append(row_data)
                except:
                    break
            
            wb.close()
            
            # TreeView에 표시
            self._update_preview_display(headers, data_rows)
            
            # 파일 정보 및 헤더 검증
            self._validate_headers_and_update_info(headers)
            
        except Exception as e:
            self._show_preview_message("미리보기를 불러올 수 없습니다.")
            logger.error(f"미리보기 로드 오류: {e}")
    
    def _update_preview_display(self, headers, data_rows):
        """미리보기 TreeView 업데이트 - 행 번호 포함"""
        try:
            # TreeView 초기화
            for item in self.preview_tree.get_children():
                self.preview_tree.delete(item)
            
            if not headers:
                self._show_preview_message("데이터가 없습니다.")
                return
            
            # 행 번호 컬럼 + 데이터 컬럼 설정 (최대 9개 데이터 컬럼 + 행번호 1개)
            display_headers = ["행번호"] + headers[:9]
            self.preview_tree['columns'] = display_headers
            
            # 헤더 설정
            self.preview_tree.heading("행번호", text="행#")
            self.preview_tree.column("행번호", width=50, minwidth=40, anchor='center')
            
            for col in headers[:9]:
                self.preview_tree.heading(col, text=col[:20])  # 헤더명 길이 제한
                self.preview_tree.column(col, width=100, minwidth=50)
            
            # 데이터 삽입 (행 번호 포함)
            for idx, row_data in enumerate(data_rows, start=2):  # 2행부터 시작 (1행은 헤더)
                # 컬럼 수에 맞춰 데이터 조정 (행번호 제외하고 9개)
                display_data = row_data[:9]
                while len(display_data) < 9:
                    display_data.append("")
                
                # 행 번호를 맨 앞에 추가
                final_data = [str(idx)] + display_data
                self.preview_tree.insert('', 'end', values=final_data)
            
            # 컬럼이 9개보다 많으면 알림
            if len(headers) > 9:
                self.preview_tree.insert('', 'end', values=["...", "..."] + ["..."] * 8)
                
        except Exception as e:
            self._show_preview_message("미리보기 표시 오류")
            logger.error(f"미리보기 표시 오류: {e}")
    
    def _validate_headers_and_update_info(self, headers):
        """헤더 검증 및 파일 정보 업데이트"""
        try:
            file_info = []
            file_info.append(f"📁 파일: {Path(self.selected_file).name}")
            file_info.append(f"📊 시트: {self.selected_sheet}")
            file_info.append(f"🚀 데이터 시작 행: {self.start_row}행")
            
            if headers:
                file_info.append(f"📋 총 {len(headers)}개 컬럼: {', '.join(headers[:5])}")
                if len(headers) > 5:
                    file_info[-1] += " ..."
                
                # 필수 헤더 검증
                has_address = "주소" in headers
                has_owner = "소유자명" in headers
                
                file_info.append("\n[헤더 검증 결과]")
                
                if has_address:
                    address_col = headers.index("주소") + 1
                    file_info.append(f"✅ '주소' 컬럼 확인됨 ({address_col}번 열)")
                else:
                    file_info.append("❌ '주소' 컬럼이 없습니다!")
                
                if has_owner:
                    owner_col = headers.index("소유자명") + 1
                    file_info.append(f"✅ '소유자명' 컬럼 확인됨 ({owner_col}번 열)")
                else:
                    file_info.append("❌ '소유자명' 컬럼이 없습니다!")
                
                # 확인 버튼 활성화 조건
                if has_address and has_owner:
                    self.ok_button['state'] = "normal"
                    self.status_label.config(text="✅ 모든 검증 완료! 자동화를 시작할 수 있습니다.", fg="green")
                else:
                    self.ok_button['state'] = "disabled"
                    missing = []
                    if not has_address:
                        missing.append("'주소'")
                    if not has_owner:
                        missing.append("'소유자명'")
                    self.status_label.config(text=f"❌ 필수 헤더가 없습니다: {', '.join(missing)}", fg="red")
            
            # 정보 표시
            self.info_text.config(state="normal")
            self.info_text.delete("1.0", "end")
            self.info_text.insert("1.0", "\n".join(file_info))
            self.info_text.config(state="disabled")
            
        except Exception as e:
            logger.error(f"헤더 검증 오류: {e}")
    
    def _confirm(self):
        """확인"""
        if self.selected_file and self.selected_sheet:
            self.start_row = self.start_row_var.get()
            self.root.quit()
        else:
            messagebox.showwarning("경고", "파일과 시트를 선택해주세요.")
    
    def _cancel(self):
        """취소"""
        self.selected_file = None
        self.selected_sheet = None
        self.start_row = None
        self.root.quit()
    
    def _on_closing(self):
        """창 닫기"""
        try:
            geometry = self.root.geometry().split('+')[0]
            self.config["window_geometry"] = geometry
            self._save_config()
        except:
            pass
        self._cancel()
    
    def show(self) -> Tuple[Optional[str], Optional[str], Optional[int]]:
        """GUI 표시"""
        try:
            self.root.mainloop()
            self.root.destroy()
            return self.selected_file, self.selected_sheet, self.start_row
        except Exception as e:
            logger.error(f"GUI 오류: {e}")
            return None, None, None


def select_excel_file_and_sheet() -> Tuple[Optional[str], Optional[str], Optional[int]]:
    """초고속 엑셀 선택기 + 미리보기 + 시작행 선택"""
    try:
        print("GUI 창 초기화 중...")
        selector = FastExcelSelector()
        print("GUI 창이 열렸습니다. 파일을 선택해주세요.")
        result = selector.show()
        print("GUI 선택 완료")
        return result
    except Exception as e:
        print(f"GUI 오류: {e}")
        return None, None, None


if __name__ == "__main__":
    file_path, sheet_name, start_row = select_excel_file_and_sheet()
    
    if file_path and sheet_name and start_row:
        print(f"선택된 파일: {file_path}")
        print(f"선택된 시트: {sheet_name}")
        print(f"시작 행: {start_row}")
    else:
        print("선택 취소됨")