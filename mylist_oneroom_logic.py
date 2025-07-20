# mylist_oneroom_logic.py
import os
import sys
import glob
import requests
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import time
import json
import logging  # 로깅 임포트 추가

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt, QModelIndex, QUrl, Q_ARG
from PyQt5.QtGui import QStandardItemModel, QStandardItem, QPixmap, QIcon, QColor, QFont
from PyQt5.QtWidgets import (
    QDialog,QTableView, QWidget, QPushButton, QVBoxLayout, QHBoxLayout, QMenu,
    QMessageBox, QAbstractItemView, QHeaderView, QInputDialog, QShortcut, QLabel
)

from ui_utils import restore_qtableview_column_widths, save_qtableview_column_widths
from dialogs import ImageSlideshowWindow, StatusChangeDialog # Assuming these exist

class MyListOneroomLogic:
    def __init__(self, parent_app, container):
        self.parent_app = parent_app
        self.container = container
        self.server_host = parent_app.server_host
        self.server_port = parent_app.server_port
        self.current_manager = parent_app.current_manager
        self.current_role = parent_app.current_role
        
        # 로거 설정 추가
        self.logger = logging.getLogger(__name__)

        # UI Elements
        self.mylist_oneroom_view = None
        self.mylist_oneroom_model = None

        # State
        self.mylist_oneroom_loading = False

        # Tab Widget container
        self.tab_widget = None

        # Define column mapping (Display Name -> DB Field Name)
        self.parent_app.COLUMN_MAP_MYLIST_ONEROOM_DISPLAY_TO_DB = {
            "주소": None, # Handled separately
            "호": "ho",
            "층": None, # Combined
            "보증금/월세": None, # Combined
            "관리비": "manage_fee",
            "입주가능일": "in_date",
            "비밀번호": "password",
            "방/화장실": None, # Combined
            "연락처": "owner_phone",
            "매물번호": None, # Combined
            "옵션": "options",
            "담당자": "manager",
            "메모": "memo",
            "주차대수": "parking",
            "용도": "building_usage",
            "사용승인일": "approval_date",
            "평수": "area",
            "광고종료일": "ad_end_date",
            "사진경로": "photo_path",
            "소유자명": "owner_name",
            "관계": "owner_relation",
            "re_ad_yn": "N", # Default new ads from Naver as "새광고"
            "status_cd": "" # Default status
        }


    def init_ui(self):
        """Creates the widget for the '원룸(새광고)' tab."""
        container_oneroom = QWidget()
        vlay_oneroom = QVBoxLayout(container_oneroom)

        # Top Buttons
        tool_layout_oneroom = QHBoxLayout()
        btn_add_oneroom = QPushButton("행 추가(원룸)")
        btn_add_oneroom.setFixedHeight(30)
        btn_add_oneroom.clicked.connect(self.on_add_mylist_oneroom_row)
        tool_layout_oneroom.addWidget(btn_add_oneroom)

        btn_save_oneroom = QPushButton("저장(원룸)")
        btn_save_oneroom.setFixedHeight(30)
        btn_save_oneroom.clicked.connect(self.on_save_mylist_oneroom_changes)
        tool_layout_oneroom.addWidget(btn_save_oneroom)

        btn_export_oneroom = QPushButton("엑셀 다운로드(원룸)")
        btn_export_oneroom.setFixedHeight(30)
        btn_export_oneroom.clicked.connect(self.export_selected_oneroom_to_excel)
        tool_layout_oneroom.addWidget(btn_export_oneroom)
        vlay_oneroom.addLayout(tool_layout_oneroom)

        # Model
        self.mylist_oneroom_model = QStandardItemModel()
        headers_oneroom = [ # Keep consistent with column map keys if possible
            "주소", "호", "층", "보증금/월세", "관리비", "입주가능일", "비밀번호",
            "방/화장실", "연락처", "매물번호", "옵션", "담당자", "메모",
            "주차대수", "용도", "사용승인일", "평수", "광고종료일", "사진경로",
            "소유자명", "관계" # "재광고" column removed as per original code? Add if needed.
        ]
        # Adjust header list if '재광고' is indeed needed for Oneroom
        # headers_oneroom.append("재광고")

        self.mylist_oneroom_model.setColumnCount(len(headers_oneroom))
        self.mylist_oneroom_model.setHorizontalHeaderLabels(headers_oneroom)
        self.mylist_oneroom_model.itemChanged.connect(self.on_mylist_oneroom_item_changed)

        # View
        self.mylist_oneroom_view = QTableView()
        self.mylist_oneroom_view.setModel(self.mylist_oneroom_model)
        self.mylist_oneroom_view.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.mylist_oneroom_view.setSortingEnabled(False) # Disable sorting

        # Restore/Save column widths
        restore_qtableview_column_widths(
            self.parent_app.settings_manager, self.mylist_oneroom_view, "MyListOneroomTable"
        )
        self.mylist_oneroom_view.horizontalHeader().sectionResized.connect(
            lambda: save_qtableview_column_widths(
                self.parent_app.settings_manager, self.mylist_oneroom_view, "MyListOneroomTable"
            )
        )

        self.mylist_oneroom_view.setSelectionMode(QAbstractItemView.ExtendedSelection) # Allow multi-row select like shop
        self.mylist_oneroom_view.setSelectionBehavior(QAbstractItemView.SelectItems)

        # Connect signals
        sel_model_oneroom = self.mylist_oneroom_view.selectionModel()
        sel_model_oneroom.currentChanged.connect(self.on_mylist_oneroom_current_changed)
        self.mylist_oneroom_view.doubleClicked.connect(self.on_mylist_oneroom_view_double_clicked)
        self.mylist_oneroom_view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.mylist_oneroom_view.customContextMenuRequested.connect(self._mylist_oneroom_context_menu)
        self.mylist_oneroom_view.horizontalHeader().sectionClicked.connect(self.on_oneroom_header_section_clicked)

        # Add Delete shortcut like the shop tab
        self.delete_shortcut_oneroom = QShortcut(QtGui.QKeySequence("Delete"), self.mylist_oneroom_view)
        self.delete_shortcut_oneroom.activated.connect(lambda: self.clear_selected_cells(self.mylist_oneroom_view))


        vlay_oneroom.addWidget(self.mylist_oneroom_view)
        container_oneroom.setLayout(vlay_oneroom)

        self.tab_widget = container_oneroom
        return self.tab_widget

    # --- Data Loading and Population ---

    def load_data(self):
        """Initiates background loading for oneroom data."""
        self.mylist_oneroom_loading = True
        future = self.parent_app.executor.submit(
            self._bg_load_mylist_oneroom_data,
            self.current_manager,
            self.current_role
        )
        future.add_done_callback(self._on_mylist_oneroom_data_fetched)

    def _bg_load_mylist_oneroom_data(self, manager, role):
        """ (Background Thread) Fetches mylist_oneroom data. """
        url = f"http://{self.server_host}:{self.server_port}/mylist/get_mylist_oneroom_data"
        params = {"manager": manager, "role": role}
        try:
            resp = requests.get(url, params=params, timeout=10)
            resp.raise_for_status()
            j = resp.json()
            if j.get("status") != "ok":
                return {"status": "error", "data": [], "message": j.get("message")}
            return {"status": "ok", "data": j.get("data", [])}
        except requests.Timeout:
             return {"status": "exception", "message": "Request timed out", "data": []}
        except requests.RequestException as ex:
            return {"status": "exception", "message": str(ex), "data": []}
        except Exception as ex_other:
             return {"status": "exception", "message": f"Unexpected error: {ex_other}", "data": []}

    def _on_mylist_oneroom_data_fetched(self, future):
        """
        (Main Thread) Processes fetched oneroom data and updates the model.
        Similar logic to the shop tab fetch callback.
        """
        self.mylist_oneroom_loading = True # Keep loading flag
        model = self.mylist_oneroom_model
        if not model:
            self.mylist_oneroom_loading = False
            return

        try:
            result = future.result()
        except Exception as e:
            QMessageBox.warning(self.parent_app, "로딩 실패", f"마이리스트(원룸) 데이터 로딩 중 오류 발생:\n{e}")
            result = None

        rows_processed, rows_added, rows_updated, rows_removed = 0, 0, 0, 0

        if result and result.get("status") == "ok":
            new_rows_from_server = result.get("data", [])
            rows_processed = len(new_rows_from_server)

            current_known_ids = self._get_mylist_oneroom_known_ids()
            pending_deleted_ids = set(self.container.pending_manager.oneroom_pending.get("deleted", []))
            if pending_deleted_ids:
                print(f"[DEBUG] MyListOneroomLogic Fetch Callback: Pending deletions to filter out: {pending_deleted_ids}")

            # Filter out pending deletions
            filtered_rows = [row for row in new_rows_from_server if row.get("id") not in pending_deleted_ids]
            filtered_count_num = rows_processed - len(filtered_rows)
            if filtered_count_num > 0:
                print(f"[DEBUG] MyListOneroomLogic Fetch Callback: Filtered out {filtered_count_num} pending deleted rows.")

            # rows to add/update
            rows_to_add = []
            rows_to_update_map = {}
            for row in filtered_rows:
                pk_id = row.get("id")
                if pk_id is None: continue
                if pk_id not in current_known_ids:
                    rows_to_add.append(row)
                else:
                    rows_to_update_map[pk_id] = row

            # Identify rows to remove
            fetched_ids = set(r.get("id") for r in filtered_rows if r.get("id") is not None)
            ids_to_remove = current_known_ids - fetched_ids
            pending_temp_ids = set(p_add.get("temp_id") for p_add in self.container.pending_manager.oneroom_pending.get("added", []))
            ids_currently_in_model_marked_as_pending_add = set()
            for r in range(model.rowCount()):
                 item_0 = model.item(r, 0)
                 if item_0:
                     item_id = item_0.data(Qt.UserRole + 3)
                     if item_id in pending_temp_ids:
                         ids_currently_in_model_marked_as_pending_add.add(item_id)

            ids_to_remove = ids_to_remove - ids_currently_in_model_marked_as_pending_add
            ids_to_remove = ids_to_remove - pending_deleted_ids

            # Apply changes to model
            try:
                self.mylist_oneroom_view.setSortingEnabled(False) # Correct

                # Remove
                if ids_to_remove:
                    indices_to_remove = []
                    for r in range(model.rowCount() - 1, -1, -1):
                         item0 = model.item(r, 0)
                         if item0 and item0.data(Qt.UserRole+3) in ids_to_remove:
                             indices_to_remove.append(r)
                    for row_idx in indices_to_remove:
                        model.removeRow(row_idx)
                        rows_removed += 1

                # Add
                if rows_to_add:
                    self.append_mylist_oneroom_rows(rows_to_add, model=model)
                    rows_added = len(rows_to_add)

                # Update
                if rows_to_update_map:
                    headers = [model.horizontalHeaderItem(j).text() for j in range(model.columnCount())] if model.columnCount() > 0 else []
                    if headers:
                        updated_in_loop = 0
                        not_found = list(rows_to_update_map.keys())
                        for r in range(model.rowCount()):
                            item0 = model.item(r, 0)
                            if not item0: continue
                            pk_id = item0.data(Qt.UserRole + 3)
                            if pk_id in rows_to_update_map:
                                row_data = rows_to_update_map[pk_id]
                                self._update_oneroom_model_row(model, r, headers, row_data) # Use helper
                                updated_in_loop += 1
                                if pk_id in not_found: not_found.remove(pk_id)
                        rows_updated = updated_in_loop
                        if not_found:
                            print(f"[WARN] MyListOneroomLogic Fetch Callback: Could not find rows to update for IDs: {not_found}")
                    else:
                        print("[WARN] MyListOneroomLogic Fetch Callback: Cannot update, model headers missing.")
            except Exception as update_err:
                 print(f"[ERROR] MyListOneroomLogic Fetch Callback: Error during model update: {update_err}")
                 QMessageBox.critical(self.parent_app, "UI 업데이트 오류", f"원룸 테이블 업데이트 중 오류 발생:\n{update_err}")
            finally:
                # --- REMOVED: Do not re-enable sorting here --- 
                # if self.mylist_oneroom_view: self.mylist_oneroom_view.setSortingEnabled(True) 
                pass # Keep finally block structure if other cleanup is needed later

        else:
            # Handle error
            error_message = "알 수 없는 상태"
            if result and result.get("status") != "ok": error_message = result.get("message", "서버 오류")
            elif not result: error_message = "서버 응답 없음"
            print(f"[ERR] MyListOneroomLogic Fetch Callback: Load failed - {error_message}")
            # self.parent_app.statusBar().showMessage(f"마이리스트(원룸) 로딩 실패: {error_message}", 5000)

        # Final steps
        self.mylist_oneroom_loading = False
        print(f"[INFO] MyListOneroomLogic data fetch/update complete. Added: {rows_added}, Updated: {rows_updated}, Removed: {rows_removed}")
        status_message = "마이리스트(원룸) 로딩 완료."
        if result and result.get("status") != "ok": status_message = f"마이리스트(원룸) 로딩 실패: {result.get('message', '')}"
        elif not result: status_message = "마이리스트(원룸) 로딩 응답 없음."
        self.parent_app.statusBar().showMessage(status_message, 3000)

    def _get_mylist_oneroom_known_ids(self):
        """Returns a set of real DB IDs currently in the oneroom model."""
        s = set()
        m = self.mylist_oneroom_model
        if not m: return s
        for r in range(m.rowCount()):
            item0 = m.item(r, 0)
            if not item0: continue
            real_id = item0.data(QtCore.Qt.UserRole+3)
            if isinstance(real_id, int) and real_id > 0:
                s.add(real_id)
        return s

    def populate_mylist_oneroom_table(self, rows):
        """Populates the oneroom table model, clearing existing data."""
        self.mylist_oneroom_loading = True
        model = self.mylist_oneroom_model
        if not model:
            self.mylist_oneroom_loading = False
            return

        try:
             model.setSortingEnabled(False)
             model.removeRows(0, model.rowCount())
             if rows:
                 self.append_mylist_oneroom_rows(rows, model=model)
        except Exception as e:
             QMessageBox.critical(self.parent_app, "테이블 채우기 오류", f"원룸 목록을 채우는 중 오류 발생:\n{e}")
        finally:
             model.setSortingEnabled(True)
             self.mylist_oneroom_loading = False
             print("[DEBUG] MyListOneroomLogic: Population complete.")


    def append_mylist_oneroom_rows(self, row_list, model=None):
        """Appends rows to the oneroom table model."""
        if not row_list: return

        m = model if model else self.mylist_oneroom_model
        if not m:
             return

        append_start_time = time.time()
        rows_to_add = len(row_list)
        start_row = m.rowCount()
        try:
            m.insertRows(start_row, rows_to_add)
            headers = [m.horizontalHeaderItem(j).text() for j in range(m.columnCount())] if m.columnCount() > 0 else []
            if not headers:
                 print("[WARN] MyListOneroomLogic: Cannot append effectively, model headers missing.")
                 return

            for i, db_row_data in enumerate(row_list):
                 row_idx = start_row + i
                 self._update_oneroom_model_row(m, row_idx, headers, db_row_data) # Use helper

        except Exception as e:
             print(f"[ERROR] MyListOneroomLogic: Error during append_mylist_oneroom_rows: {e}")
        finally:
             append_duration = time.time() - append_start_time

    def _update_oneroom_model_row(self, model, row_idx, headers, db_row_data):
        """ Helper function to set items for a single row in the oneroom model. """
        item0 = None
        for col_idx, header_name in enumerate(headers):
            db_key = self.parent_app.COLUMN_MAP_MYLIST_ONEROOM_DISPLAY_TO_DB.get(header_name, None)
            raw_value = db_row_data.get(db_key) if db_key else None

            # Format cell value based on header
            if header_name == "주소": cell_val = f"{db_row_data.get('dong', '')} {db_row_data.get('jibun', '')}".strip()
            elif header_name == "층": cell_val = f"{db_row_data.get('curr_floor', 0)}/{db_row_data.get('total_floor', 0)}"
            elif header_name == "보증금/월세": cell_val = f"{db_row_data.get('deposit', 0)}/{db_row_data.get('monthly', 0)}"
            elif header_name == "방/화장실": cell_val = f"{db_row_data.get('rooms', 0)}/{db_row_data.get('baths', 0)}"
            elif header_name == "매물번호": cell_val = f"{db_row_data.get('naver_property_no', '')}/{db_row_data.get('serve_property_no', '')}"
            elif header_name == "평수": cell_val = str(raw_value) if raw_value is not None else ""
            elif header_name == "관리비": cell_val = str(raw_value) if raw_value is not None else ""
            elif header_name == "주차대수": cell_val = str(raw_value) if raw_value is not None else ""
            elif header_name == "사용승인일": cell_val = str(raw_value) if raw_value is not None else ""
            elif header_name == "광고종료일": cell_val = str(raw_value) if raw_value is not None else ""
            # Add 재광고 if needed:
            # elif header_name == "재광고": cell_val = "재광고" if db_row_data.get("re_ad_yn", "N") == "Y" else "새광고"
            else: cell_val = str(raw_value) if raw_value is not None else ""

            item = self._create_oneroom_item(header_name, cell_val, db_row_data)
            model.setItem(row_idx, col_idx, item)
            if col_idx == 0: item0 = item

        # Optional: Set row background color if needed (e.g., based on re_ad_yn)
        # re_ad_yn = db_row_data.get("re_ad_yn", "N") == "Y"
        # row_bg = QColor("#E1BEE7") if re_ad_yn else QColor("#FFFFFF") # Example: purple or white
        # for c in range(model.columnCount()):
        #     cell = model.item(row_idx, c)
        #     if cell: cell.setBackground(row_bg)

    def _create_oneroom_item(self, header_name, cell_value, db_row_data):
        """ Creates a QStandardItem for the oneroom table. """
        item = QStandardItem(str(cell_value))

        if header_name == "주소":
            folder_path = db_row_data.get("photo_path", "") or ""
            rep_img_path = ""
            if folder_path and os.path.isdir(folder_path):
                 try:
                    files = [f for f in os.listdir(folder_path) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.gif'))]
                    if files: rep_img_path = os.path.join(folder_path, sorted(files)[0])
                 except OSError as e: print(f"[WARN] Cannot access folder path '{folder_path}': {e}")

            if rep_img_path and os.path.isfile(rep_img_path):
                try:
                    pixmap = QPixmap(rep_img_path).scaled(24, 24, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    if not pixmap.isNull():
                        item.setIcon(QIcon(pixmap))
                        file_url = QUrl.fromLocalFile(rep_img_path).toString()
                        item.setToolTip(f'<img src="{file_url}" width="200">')
                    else: item.setToolTip(cell_value)
                except Exception as img_err:
                     print(f"[WARN] Error creating oneroom icon/tooltip for {rep_img_path}: {img_err}")
                     item.setToolTip(cell_value)
            else: item.setToolTip(cell_value)

            item.setData(folder_path, Qt.UserRole + 10)
            item.setData(rep_img_path, Qt.UserRole + 11)
            item.setData(db_row_data.get("status_cd", ""), Qt.UserRole + 1)
            item.setData(db_row_data.get("id"), Qt.UserRole + 3)

        return item

    # --- Item Changed Signal ---

    def on_mylist_oneroom_item_changed(self, item: QStandardItem):
        """Handles item changes in the oneroom table model."""
        if self.mylist_oneroom_loading: return # Ignore changes during load
        
        row = item.row(); col = item.column()
        model = self.mylist_oneroom_model
        item0 = model.item(row, 0)
        if not item0: return
        pk_id = item0.data(Qt.UserRole + 3)
        if pk_id is None: return

        # --- 변경 시작: PendingManager 사용 ---
        header_item = model.horizontalHeaderItem(col)
        if not header_item: return
        header_text = header_item.text()
        new_value = item.text()
        
        if pk_id > 0: # Only track updates for existing DB rows
            # Parse the row to get all fields for update payload
            try:
                 parsed_row_data = self._parse_mylist_oneroom_row(row)
                 if not parsed_row_data: # Parsing failed
                      self.logger.warning(f"Failed to parse oneroom row {row} on item change, cannot add pending update.")
                      return

                 update_payload = {"id": pk_id}
                 # Add only the changed field(s) based on header_text
                 # Need a mapping similar to shop logic or enhance parsing
                 # For now, let's assume _parse_mylist_oneroom_row gives DB fields
                 # and we send the *entire* parsed row data for simplicity 
                 # (server side should handle partial updates ideally)
                 # NOTE: This might overwrite other pending changes for the same row ID
                 # if multiple cells are edited before save.
                 # A more robust approach would be to get the specific DB field 
                 # corresponding to header_text and only add that field to the payload.
                 update_payload.update(parsed_row_data) # Update with all parsed fields
                 
                 # Add to pending updates using the manager
                 self.container.pending_manager.add_pending_oneroom_update(update_payload)
                 self.logger.debug(f"Added/Updated pending oneroom update via manager for ID {pk_id}")
            except Exception as e:
                 self.logger.error(f"Error parsing oneroom row {row} or adding pending update: {e}", exc_info=True)
        else:
             self.logger.debug(f"Item changed in temporary oneroom row (temp_id={pk_id}). Update handled during save.")
        # --- 변경 끝 ---

    # --- Context Menu and Actions ---

    def _mylist_oneroom_context_menu(self, pos):
        """Creates and shows the context menu for the oneroom table."""
        index = self.mylist_oneroom_view.indexAt(pos)
        if not index.isValid(): return

        menu = QMenu(self.mylist_oneroom_view)
        act_copy = menu.addAction("행 복사(원룸)")
        act_delete = menu.addAction("행 삭제(원룸)")
        act_completed_oneroom = menu.addAction("상태 변경(계약완료/기타)")

        action = menu.exec_(self.mylist_oneroom_view.mapToGlobal(pos))

        if action == act_copy:
            self.copy_mylist_oneroom_row(index.row())
        elif action == act_delete:
            self.delete_selected_mylist_oneroom_rows()
        elif action == act_completed_oneroom:
             self.on_mylist_oneroom_change_status(pos)

    def copy_mylist_oneroom_row(self, source_row_idx):
        """Copies a row and adds it as a new temporary oneroom row."""
        model = self.mylist_oneroom_model
        if not model or source_row_idx < 0 or source_row_idx >= model.rowCount(): return
        col_count = model.columnCount()
        copied_values = [model.item(source_row_idx, c).text() if model.item(source_row_idx, c) else "" for c in range(col_count)]
        # Use container's method to add row (already uses pending manager)
        self.container.add_new_oneroom_row(initial_data=copied_values)

    def delete_selected_mylist_oneroom_rows(self):
        """선택된 행을 '삭제 예정' 상태로 표시하고, pending changes에 기록합니다."""
        view = self.mylist_oneroom_view
        model = self.mylist_oneroom_model
        sel_model = view.selectionModel()
        # --- 변경: container 대신 pending_manager 사용 ---
        pending_manager = self.container.pending_manager
        
        if not view or not model or not sel_model or not pending_manager: # Check pending_manager
             self.logger.error("delete_selected_mylist_oneroom_rows: View, model, selection model, or pending_manager is None.")
             return
        # --- 변경 끝 ---
        
        selected_indexes = sel_model.selectedIndexes()
        if not selected_indexes: return
        involved_rows = set(idx.row() for idx in selected_indexes)
        if not involved_rows: return

        reply = QMessageBox.question(view, "삭제 확인", f"선택한 셀이 포함된 {len(involved_rows)}개 행 전체를 삭제 상태로 표시하시겠습니까?",
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.No: return

        marked_count = 0
        rows_to_mark = sorted(list(involved_rows))
        for row_idx in rows_to_mark:
            item0 = model.item(row_idx, 0)
            record_id = None
            if item0:
                record_id = item0.data(Qt.UserRole + 3)
            
            # Mark row visually
            self.mark_row_as_pending_deletion(model, row_idx)
            marked_count += 1
            
            if record_id is not None:
                # --- 변경: Mark for deletion using pending_manager ---
                pending_manager.mark_oneroom_row_for_deletion(record_id)
                self.logger.debug(f"Marked oneroom row for deletion via pending manager: ID={record_id}")
                # ---------------------------------------------------

        self.logger.info(f"Marked {marked_count} oneroom rows for deletion. Pending deletes: {pending_manager.get_pending_oneroom_changes()['deleted']}")

    def on_mylist_oneroom_change_status(self, pos):
        """Handles the '상태 변경' action for the oneroom context menu."""
        view = self.mylist_oneroom_view
        sel_model = view.selectionModel()

        selected_rows_indices = []
        if sel_model and sel_model.hasSelection():
            selected_rows_indices = sorted(list(set(idx.row() for idx in sel_model.selectedIndexes())))
        else:
            index_under_mouse = view.indexAt(pos)
            if index_under_mouse.isValid():
                selected_rows_indices = [index_under_mouse.row()]

        if not selected_rows_indices: return

        if self.current_role == "admin":
            all_managers = [self.parent_app.manager_dropdown.itemText(i) for i in range(self.parent_app.manager_dropdown.count())]
        else:
            all_managers = [self.current_manager]

        dlg = StatusChangeDialog(
            current_role=self.current_role, current_manager=self.current_manager,
            all_managers=all_managers, parent=self.parent_app
        )
        if dlg.exec_() != QDialog.Accepted: return

        result = dlg.get_result()
        chosen_manager = result["manager"]
        chosen_status = result["status"]
        memo_text = result["memo"]

        model = self.mylist_oneroom_model
        items_arr = []
        rows_to_remove_from_ui = []
        parsed_failed_ids = []
        for row_idx in selected_rows_indices:
             try:
                 parsed_data = self._parse_mylist_oneroom_row(row_idx)
                 if parsed_data:
                     item_id = parsed_data.get("id")
                     if item_id is not None and item_id > 0:
                         parsed_data["manager"] = chosen_manager
                         parsed_data["status_cd"] = chosen_status
                         parsed_data["memo"] = memo_text
                         items_arr.append(parsed_data)
                         rows_to_remove_from_ui.append(row_idx)
                     else:
                          parsed_failed_ids.append(f"Row {row_idx} (Temp ID or missing)")
                 else:
                      parsed_failed_ids.append(f"Row {row_idx} (Parsing failed)")
             except Exception as e:
                 self.logger.error(f"Error parsing oneroom row {row_idx} for status change: {e}", exc_info=True)
                 parsed_failed_ids.append(f"Row {row_idx} (Exception: {e})")
        
        if parsed_failed_ids:
             QMessageBox.warning(view, "일부 처리 불가", f"다음 행은 상태 변경 처리 중 문제가 발생했습니다:\n" + "\n".join(parsed_failed_ids))

        if not items_arr:
             QMessageBox.information(view, "변경 없음", "상태를 변경할 유효한 항목이 없습니다.")
             return

        payload = {
            "deals": items_arr,
            "manager": self.current_manager,
            "role": self.current_role
        }
        # Submit task via container (status change logic remains in container for now)
        self.container.submit_status_change_task(payload, rows_to_remove_from_ui, 'oneroom')

    # --- Manual Save Action --- 

    def on_save_mylist_oneroom_changes(self):
        """Triggers the manual save process for the oneroom tab via the save handler."""
        # --- 변경: container 대신 save_handler 사용 ---
        if self.container and self.container.save_handler:
            self.logger.info("Manual save requested for oneroom tab. Forwarding to save_handler.")
            self.container.save_handler.save_pending_oneroom_changes()
        else:
            self.logger.error("Cannot save oneroom changes: container or save_handler is missing.")
            QMessageBox.warning(self.parent_app, "오류", "저장 핸들러를 찾을 수 없어 저장할 수 없습니다.")
        # --- 변경 끝 ---

    # --- Saving Changes ---

    def _build_mylist_oneroom_rows_for_changes(self, added_list, updated_list):
        """Parses oneroom rows from the model based on pending IDs."""
        model = self.mylist_oneroom_model
        if not model: return {"added": [], "updated": []}

        rowCount = model.rowCount()
        id_to_row = {model.item(r, 0).data(Qt.UserRole+3): r for r in range(rowCount) if model.item(r, 0) and model.item(r, 0).data(Qt.UserRole+3) is not None}

        added_rows_parsed = []
        for add_obj in added_list:
            temp_id = add_obj.get("temp_id")
            row_idx = id_to_row.get(temp_id)
            if row_idx is not None:
                row_dict = self._parse_mylist_oneroom_row(row_idx)
                row_dict["temp_id"] = temp_id
                added_rows_parsed.append(row_dict)

        updated_rows_parsed = []
        updated_ids = set(upd_obj.get("id") for upd_obj in updated_list if isinstance(upd_obj.get("id"), int) and upd_obj.get("id") > 0)
        for real_id in updated_ids:
            row_idx = id_to_row.get(real_id)
            if row_idx is not None:
                row_dict = self._parse_mylist_oneroom_row(row_idx)
                row_dict["id"] = real_id
                updated_rows_parsed.append(row_dict)

        return {"added": added_rows_parsed, "updated": updated_rows_parsed}


    def _parse_mylist_oneroom_row(self, row_idx):
        """Parses a single row from the oneroom model into a DB-ready dictionary."""
        model = self.mylist_oneroom_model
        if not model or row_idx >= model.rowCount(): return {}

        def get_item_text(r, c):
            item = model.item(r, c)
            return item.text().strip() if item else ""

        # Parse based on known column indices/headers
        addr_text = get_item_text(row_idx, 0); dong_val, jb_val = addr_text.split(" ", 1) if " " in addr_text else (addr_text, "")
        ho_val = get_item_text(row_idx, 1)
        fl_text = get_item_text(row_idx, 2); cf, tf = (int(p) if p.isdigit() else 0 for p in fl_text.split("/", 1)) if "/" in fl_text else (0,0)
        bm_text = get_item_text(row_idx, 3); dp, mn = (int(p) if p.isdigit() else 0 for p in bm_text.split("/", 1)) if "/" in bm_text else (0,0)
        mg_val = get_item_text(row_idx, 4) # 관리비
        ind_val = get_item_text(row_idx, 5) # 입주가능일
        pwd_val = get_item_text(row_idx, 6) # 비밀번호
        rb_text = get_item_text(row_idx, 7); ro, ba = rb_text.split("/", 1) if "/" in rb_text else ("0","0")
        ph_val = get_item_text(row_idx, 8) # 연락처
        mm_text = get_item_text(row_idx, 9); nav, srv = mm_text.split("/", 1) if "/" in mm_text else (mm_text, "")
        opt_val = get_item_text(row_idx, 10) # 옵션
        mgr_val = get_item_text(row_idx, 11) # 담당자
        memo_val = get_item_text(row_idx, 12) # 메모
        pk_val = get_item_text(row_idx, 13) # 주차대수
        bu_val = get_item_text(row_idx, 14) # 용도
        ap_val = get_item_text(row_idx, 15) # 사용승인일
        ar_text = get_item_text(row_idx, 16)
        ar_val = 0.0
        try:
            ar_val = float(ar_text) if ar_text else 0.0
        except ValueError:
            ar_val = 0.0

        ae_val = get_item_text(row_idx, 17) # 광고종료일
        pt_val = get_item_text(row_idx, 18) # 사진경로
        on_val = get_item_text(row_idx, 19) # 소유자명
        or_val = get_item_text(row_idx, 20) # 관계

        status_cd = ""
        item0 = model.item(row_idx, 0)
        if item0: status_cd = item0.data(Qt.UserRole + 1) or ""

        # Construct dict matching DB fields expected by `/update_mylist_oneroom_items`
        row_dict = {
            "gu": "", "dong": dong_val, "jibun": jb_val, "ho": ho_val,
            "curr_floor": cf, "total_floor": tf, "deposit": dp, "monthly": mn,
            "manage_fee": mg_val, "in_date": ind_val, "password": pwd_val,
            "rooms": ro, "baths": ba, "options": opt_val, "owner_phone": ph_val,
            "building_usage": bu_val, "naver_property_no": nav.strip(),
            "serve_property_no": srv.strip(), "approval_date": ap_val,
            "memo": memo_val, "manager": mgr_val if mgr_val else self.current_manager, # Default to current user
            "photo_path": pt_val, "owner_name": on_val, "owner_relation": or_val,
            "ad_end_date": ae_val, "lat": 0.0, "lng": 0.0, "parking": pk_val,
            "area": ar_val, "status_cd": status_cd,
            "re_ad_yn": "N", # Default new ads from Naver as "새광고"
            "status_cd": "" # Default status
        }
        return row_dict

    def _update_mylist_oneroom_row_id(self, old_tid, new_id):
        """Updates the temporary ID to the real DB ID in the oneroom model."""
        m = self.mylist_oneroom_model
        if not m: return
        for r in range(m.rowCount()):
            item0 = m.item(r, 0)
            if not item0: continue
            rid = item0.data(Qt.UserRole+3)
            if rid == old_tid:
                item0.setData(new_id, Qt.UserRole+3)
                print(f"[INFO] MyListOneroomLogic: Updated temp_id={old_tid} to real_id={new_id} in model.")
                return

    # --- Other UI Handlers ---

    def on_add_mylist_oneroom_row(self):
        """Adds a new blank row for oneroom using the container."""
        self.container.add_new_oneroom_row()


    def on_mylist_oneroom_view_double_clicked(self, index: QModelIndex):
        """Handles double-click on the oneroom table (e.g., show images)."""
        if not index.isValid() or self.mylist_oneroom_loading: return

        if index.column() == 0: # Address column
            model = self.mylist_oneroom_model
            item_ = model.item(index.row(), 0)
            if not item_: return
            folder_path = item_.data(Qt.UserRole + 10) or ""
            if not folder_path or not os.path.isdir(folder_path): return

            try:
                image_files = sorted(
                    glob.glob(os.path.join(folder_path, "*.jpg")) +
                    glob.glob(os.path.join(folder_path, "*.jpeg")) +
                    glob.glob(os.path.join(folder_path, "*.png")) +
                    glob.glob(os.path.join(folder_path, "*.gif"))
                )
            except Exception as e:
                 return

            if not image_files: return

            # Use parent's slider window
            if hasattr(self.parent_app, 'slider_window') and self.parent_app.slider_window:
                 if self.parent_app.slider_window.isVisible():
                     self.parent_app.slider_window.set_image_list(image_files)
                     self.parent_app.slider_window.activateWindow(); self.parent_app.slider_window.raise_()
                     return
            self.parent_app.slider_window = ImageSlideshowWindow(image_files, parent=self.parent_app)
            self.parent_app.slider_window.show()

    def on_mylist_oneroom_current_changed(self, current: QModelIndex, previous: QModelIndex):
        """Handles selection change in the oneroom table."""
        if not current.isValid() or self.mylist_oneroom_loading or current.row() == previous.row():
            return

        new_row = current.row()
        addr_item = self.mylist_oneroom_model.item(new_row, 0)
        if not addr_item: return
        address_text = addr_item.text().strip()
        if not address_text: return

        # Update parent state
        if hasattr(self.parent_app, 'update_selection_from_manager_check'):
             print(f"[DEBUG] MyListOneroomLogic: Calling parent update for: {address_text}")
             self.parent_app.update_selection_from_manager_check(address_text)
        else:
             print("[ERR] MyListOneroomLogic: parent_app.update_selection_from_manager_check missing.")

    def export_selected_oneroom_to_excel(self):
        """Exports selected oneroom rows to Excel."""
        view = self.mylist_oneroom_view
        sel_model = view.selectionModel()
        if not sel_model or not sel_model.hasSelection():
             QMessageBox.information(view, "선택 없음", "엑셀로 내보낼 행을 선택하세요.")
             return

        row_indexes = sorted(list(set(idx.row() for idx in sel_model.selectedIndexes())))
        if not row_indexes: return

        model = self.mylist_oneroom_model

        # Define Excel mapping (adjust headers as needed)
        excel_map = [
            (0, "주소", "주소"), (1, "호", "호"), (2, "층", "층"), (3, "보증금/월세", "보증금/월세"),
            (4, "관리비", "관리비"), (5, "입주가능일", "입주가능일"), (6, "비밀번호", "비밀번호"),
            (7, "방/화장실", "방/화장실"), (8, "연락처", "연락처"), (9, "매물번호", "매물번호"),
            (10, "옵션", "옵션"), (11, "담당자", "담당자"), (12, "메모", "메모"),
            (13, "주차대수", "주차대수"), (14, "용도", "용도"), (15, "사용승인일", "사용승인일"),
            (16, "평수", "평수"), (17, "광고종료일", "광고종료일"), (18, "사진경로", "사진경로"),
            (19, "소유자명", "소유자명"), (20, "관계", "관계")
        ]
        model_headers = [model.horizontalHeaderItem(j).text() for j in range(model.columnCount())]

        wb = Workbook()
        ws = wb.active

        # Write headers
        for excel_col_idx, header_text, _ in excel_map:
            ws.cell(row=1, column=excel_col_idx + 1).value = header_text

        # Write data
        excel_row_num = 2
        for model_row_idx in row_indexes:
            for excel_col_idx, _, model_header_name in excel_map:
                 try:
                      model_col_idx = model_headers.index(model_header_name)
                      item_ = model.item(model_row_idx, model_col_idx)
                      text_val = item_.text().strip() if item_ else ""
                      # Add specific formatting if needed (e.g., remove hyphens from dates)
                      if model_header_name == "사용승인일":
                           text_val = text_val.replace("-", "")
                      ws.cell(row=excel_row_num, column=excel_col_idx + 1).value = text_val
                 except ValueError:
                      print(f"[WARN] Column '{model_header_name}' not found for Oneroom Excel export.")
                 except Exception as cell_err:
                      print(f"[ERROR] Error processing Oneroom cell R{model_row_idx} C'{model_header_name}': {cell_err}")
                      ws.cell(row=excel_row_num, column=excel_col_idx + 1).value = "#ERROR#"
            excel_row_num += 1

        # Save file
        try:
            now_str = datetime.now().strftime("%Y%m%d_%H%M")
            manager_str = self.current_manager.replace(" ", "_") if self.current_manager else "UnknownMgr"
            filename = f"{now_str}_{manager_str}_원룸.xlsx"
            save_path = os.path.join(os.getcwd(), filename)
            wb.save(save_path)
            QMessageBox.information(view, "엑셀 저장 완료", f"파일이 저장되었습니다:\n{save_path}")
            print(f"[INFO] Oneroom Excel export complete => {save_path}")
        except Exception as save_err:
            print(f"[ERROR] Failed to save Oneroom Excel file: {save_err}")
            QMessageBox.critical(view, "엑셀 저장 실패", f"파일 저장 중 오류 발생:\n{save_err}")

    def clear_selected_cells(self, table_view: QTableView):
        """Clears the content of the selected cells in the oneroom table."""
        sel_model = table_view.selectionModel()
        if not sel_model or not sel_model.hasSelection(): return

        model = table_view.model()
        model.blockSignals(True)
        try:
            selected_indexes = sel_model.selectedIndexes()
            changed = False
            for idx in selected_indexes:
                if idx.isValid():
                    # --- 추가: 보호할 열 지정 (주소 등) ---
                    protected_columns = [0] # 0번 열 (주소) 보호
                    if idx.column() not in protected_columns:
                        current_text = model.data(idx, Qt.DisplayRole)
                        if current_text != "":
                            model.setData(idx, "")
                            self.on_mylist_oneroom_item_changed(model.itemFromIndex(idx)) # Trigger update logic
                            changed = True
                    # ---------------------------------------
        finally:
            model.blockSignals(False)
            # No summary recalc needed for oneroom currently


    # --- Filtering ---
    def filter_table_by_address(self, address_str: str):
        """Hides or shows rows based on the address string."""
        view = self.mylist_oneroom_view
        model = self.mylist_oneroom_model
        if not view or not model: return

        try:
            model.setSortingEnabled(False)
            for row in range(model.rowCount()):
                item = model.item(row, 0) # Address column
                if item:
                    row_addr = item.text().strip()
                    should_hide = bool(address_str) and (row_addr != address_str)
                    view.setRowHidden(row, should_hide)
                else:
                    view.setRowHidden(row, bool(address_str)) # Hide if no address and filtering
        except Exception as e:
             print(f"[ERROR] MyListOneroomLogic: Error during filtering: {e}")
        finally:
             if model: model.setSortingEnabled(True) 

    # --- ADDED: Header Click Handler and Helper --- 
    def on_oneroom_header_section_clicked(self, logical_index: int):
        """Handles header clicks for the oneroom table."""
        mods = QtWidgets.QApplication.keyboardModifiers()
        if mods & QtCore.Qt.ControlModifier:
            # Ctrl + Click: Perform sorting
            current_order = self.mylist_oneroom_view.horizontalHeader().sortIndicatorOrder()
            new_order = Qt.DescendingOrder if current_order == Qt.AscendingOrder else Qt.AscendingOrder
            self.mylist_oneroom_view.sortByColumn(logical_index, new_order)
        else:
            # Normal Click: Select the entire column
            self.select_entire_column(self.mylist_oneroom_view, logical_index)

    def select_entire_column(self, table_view: QTableView, col_index: int):
        """Selects all cells in the specified column."""
        model = table_view.model()
        if not model or model.rowCount() == 0: return
        selection = QtCore.QItemSelection(model.index(0, col_index), model.index(model.rowCount() - 1, col_index))
        table_view.selectionModel().clearSelection()
        table_view.selectionModel().select(selection, QtCore.QItemSelectionModel.Select) 

    # --- ADDED METHOD ---
    def get_summary_by_manager(self, manager_name):
        """Calculates the count of listings assigned to a specific manager."""
        count = 0
        model = self.mylist_oneroom_model # Use oneroom model
        if not model:
            print("[WARN] OneroomLogic: Model not available for summary calculation.")
            return {"assigned": 0}

        manager_col_index = -1
        try:
            headers = [model.horizontalHeaderItem(j).text() for j in range(model.columnCount())]
            manager_col_index = headers.index("담당자")
        except (ValueError, AttributeError):
            print("[WARN] OneroomLogic: Cannot find '담당자' column for summary.")
            return {"assigned": 0} # Indicate error or inability to calculate

        for r in range(model.rowCount()):
            item = model.item(r, manager_col_index)
            if item and item.text().strip() == manager_name:
                count += 1

        return {"assigned": count}
    # --- END ADDED METHOD --- 

    def mark_row_as_pending_deletion(self, model, row_index):
        """지정된 행의 스타일을 '삭제 예정'으로 변경합니다."""
        if not model: return
        col_count = model.columnCount()
        for col in range(col_count):
            item = model.item(row_index, col)
            if not item:
                item = QStandardItem()
                model.setItem(row_index, col, item)
            item.setBackground(QColor("#DDDDDD")) 
            font = item.font()
            font.setStrikeOut(True)
            item.setFont(font)
            item.setData(True, Qt.UserRole + 20) # 삭제 예정 플래그 